"""
演示金额列修复效果
"""
from 数据填充工具 import DataFiller
import openpyxl
import os

def demonstrate_fix():
    """演示修复前后的对比"""
    print("\n" + "=" * 70)
    print("金额列修复演示")
    print("=" * 70)

    # 加载测试文件
    test_file = "temp/金融岛报表_交接单_20260109_144153.xlsx"

    if not os.path.exists(test_file):
        print(f"\n❌ 测试文件不存在: {test_file}")
        print("   请先将测试文件放在temp文件夹下")
        return

    print(f"\n📄 测试文件: {test_file}")

    # 执行修复后的测试
    try:
        filler = DataFiller(test_file)
        filler.load_report()

        print(f"\n✅ 加载报表成功")

        # 清空并重新填充R-W列
        ws = filler.wb['1']

        # 获取供应商数据范围
        source_start_col = 29  # AC列
        source_end_col = 34    # AH列
        target_start_col = 18  # R列
        target_end_col = 23    # W列
        target_row = 2         # 从第2行开始

        # 清空R-W列
        filler._clear_target_columns(ws, target_start_col, target_end_col)

        # 复制数据（使用修复后的逻辑）
        print(f"\n📋 使用修复后的逻辑复制数据...")
        rows_copied = filler.copy_data_to_handover(
            ws,
            source_start_col,
            source_end_col,
            target_start_col,
            target_end_col,
            target_row,
            recalculate_amount=True
        )

        print(f"✅ 复制了 {rows_copied} 行数据")

        # 保存结果
        output_file = "temp/demo_修复后.xlsx"
        filler.save_report(output_file)

        print(f"\n✅ 已保存到: {output_file}")

        # 显示前几行的金额列
        print(f"\n📊 修复后的金额列（V列）内容：")
        print(f"{'行号':<6} {'行类型':<8} {'R列内容':<20} {'V列（金额）':<30}")
        print("-" * 70)

        for row in range(2, min(11, ws.max_row + 1)):
            r_cell = ws.cell(row=row, column=18)  # R列
            v_cell = ws.cell(row=row, column=22)  # V列（金额列）

            r_value = str(r_cell.value)[:18] if r_cell.value else ""
            v_value = str(v_cell.value)[:28] if v_cell.value else ""

            # 判断行类型
            if row == 2:
                row_type = "表头"
            elif "合计" in r_value:
                row_type = "合计"
            elif r_value and r_value not in ["产品", "单位", "数量", "单价", "金额", "备注"]:
                row_type = "数据"
            else:
                row_type = "空"

            print(f"{row:<6} {row_type:<8} {r_value:<20} {v_value:<30}")

        print("\n" + "=" * 70)
        print("✅ 修复完成！")
        print("=" * 70)

        print(f"\n💡 提示：")
        print(f"   1. 打开 {output_file} 查看完整结果")
        print(f"   2. 修改数量或单价，金额会自动更新")
        print(f"   3. 合计行使用了 =SUM 公式")

    except Exception as e:
        print(f"\n❌ 演示失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    demonstrate_fix()
