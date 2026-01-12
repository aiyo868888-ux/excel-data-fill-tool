"""
创建测试Excel文件 - 验证非数字工作表是否被清空
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys

# 设置输出编码为utf-8
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

def create_test_file():
    """创建包含多种工作表的测试文件"""

    print("=" * 70)
    print("创建测试Excel文件...")
    print("=" * 70)

    wb = Workbook()

    # 删除默认工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建工作表列表
    sheets = [
        '封面',
        '目录',
        '汇总',
        '统计',
        '说明',
        '1', '2', '3', '4', '5'
    ]

    for sheet_name in sheets:
        ws = wb.create_sheet(title=sheet_name)

        # 在A列添加内容
        for row in range(1, 21):
            cell = ws.cell(row=row, column=1, value=f"{sheet_name}-A{row}")
            cell.font = Font(name='微软雅黑', size=11)

        # 在R-W列（18-23列）添加重要数据
        for row in range(1, 21):
            for col in range(18, 24):  # R到W列
                cell = ws.cell(row=row, column=col,
                             value=f"{sheet_name}-R{row}-C{col}")
                cell.font = Font(name='微软雅黑', size=10)
                cell.fill = PatternFill(start_color='FFCCCC',
                                       end_color='FFCCCC',
                                       fill_type='solid')

        print(f"   ✅ 创建工作表: {sheet_name} (包含A列和R-W列数据)")

    # 在"汇总"工作表添加一些重要统计数据
    ws_summary = wb['汇总']
    ws_summary['R2'] = "重要统计"
    ws_summary['R3'] = "总数"
    ws_summary['S3'] = 1000
    ws_summary['R4'] = "合计"
    ws_summary['S4'] = 5000
    print(f"   ✅ 汇总工作表添加了统计数据")

    # 保存文件
    filename = 'test_non_numeric_sheets.xlsx'
    wb.save(filename)

    print(f"\n✅ 测试文件已创建: {filename}")
    print(f"   文件路径: {os.path.abspath(filename)}")
    print(f"   工作表数量: {len(sheets)}")
    print(f"   数字工作表: 5个 (1-5)")
    print(f"   非数字工作表: 5个 (封面, 目录, 汇总, 统计, 说明)")

    return filename

def check_file_before(filename):
    """检查操作前的文件状态"""
    print("\n" + "=" * 70)
    print("检查操作前的文件状态...")
    print("=" * 70)

    from openpyxl import load_workbook
    wb = load_workbook(filename)

    before_state = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 检查R列第2行是否有数据
        r2_value = ws['R2'].value
        has_data = r2_value is not None and str(r2_value).strip() != ''

        before_state[sheet_name] = {
            'R2_value': r2_value,
            'has_data': has_data
        }

        print(f"   {sheet_name:8s} - R2单元格: '{r2_value}' (有数据: {has_data})")

    wb.close()
    return before_state

if __name__ == '__main__':
    # 创建测试文件
    test_file = create_test_file()

    # 检查初始状态
    before = check_file_before(test_file)

    print("\n" + "=" * 70)
    print("下一步：使用Web界面上传此文件并执行交接单操作")
    print("然后检查非数字工作表的R-W列是否被清空")
    print("=" * 70)
