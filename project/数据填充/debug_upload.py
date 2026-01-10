"""
调试上传送货商文件的问题
"""
import requests
import sys
sys.stdout.reconfigure(encoding="utf-8")

BASE_URL = "http://localhost:8888"

print("=" * 70)
print("测试上传送货商文件")
print("=" * 70)

# 步骤1：创建会话
print("\n1. 创建会话...")
resp = requests.post(f"{BASE_URL}/api/session/create")
session_id = resp.json()['session_id']
print(f"   会话ID: {session_id}")

# 步骤2：创建测试报表并上传
print("\n2. 创建测试报表...")
import openpyxl
wb = openpyxl.Workbook()
# 删除默认sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
# 创建工作表1
ws1 = wb.create_sheet("1")
ws1['A1'] = "测试"
report_file = "test_report.xlsx"
wb.save(report_file)
wb.close()
print(f"   测试报表已创建: {report_file}")

# 上传报表
print("\n3. 上传报表...")
with open(report_file, 'rb') as f:
    files = {'report': f}
    data = {'password': '1'}
    headers = {'X-Session-ID': session_id}
    resp = requests.post(f"{BASE_URL}/api/upload-report", files=files, data=data)
    print(f"   状态码: {resp.status_code}")
    print(f"   响应: {resp.text[:500]}")

# 步骤3：创建送货商文件
print("\n4. 创建送货商文件...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "9"
ws['A1'] = "名称"
ws['A2'] = "商品A"
ws['B1'] = "数量"
ws['B2'] = "10"
supplier_file = "test_supplier.xlsx"
wb.save(supplier_file)
wb.close()
print(f"   送货商文件已创建: {supplier_file}")

# 步骤4：上传送货商文件
print("\n5. 上传达货商文件...")
with open(supplier_file, 'rb') as f:
    files = {'files': f}
    headers = {'X-Session-ID': session_id}
    resp = requests.post(f"{BASE_URL}/api/upload-suppliers", files=files, headers=headers)
    print(f"   状态码: {resp.status_code}")
    print(f"   响应头: {dict(resp.headers)}")
    print(f"   响应内容类型: {resp.headers.get('content-type')}")
    print(f"   响应内容（前500字符）: {resp.text[:500]}")

    # 尝试解析JSON
    try:
        json_data = resp.json()
        print(f"   JSON数据: {json_data}")
    except Exception as e:
        print(f"   JSON解析失败: {e}")

# 清理
import os
os.remove(report_file)
os.remove(supplier_file)
print("\n6. 测试完成，临时文件已清理")
