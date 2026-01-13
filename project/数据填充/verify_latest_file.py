"""
验证最新生成的交接单文件的金额列
"""
import openpyxl
import os
import sys
import io

# 修复编码
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def verify_latest_handover():
    """验证最新生成的交接单文件"""
    print("=" * 70)
    print("验证最新生成的交接单文件")
    print("=" * 70)

    # 查找最新的交接单文件
    temp_dir = "d:/claude code -11/project/数据填充/temp"
    files = [
        f for f in os.listdir(temp_dir)
        if f.startswith("金融岛报表_交接单_") and f.endswith(".xlsx")
    ]

    if not files:
        print("\n❌ 没有找到交接单文件")
        return

    # 按修改时间排序，取最新的
    files_with_time = [
        (f, os.path.getmtime(os.path.join(temp_dir, f)))
        for f in files
    ]
    files_with_time.sort(key=lambda x: x[1], reverse=True)
    latest_file = os.path.join(temp_dir, files_with_time[0][0])

    print(f"\n📄 最新文件: {os.path.basename(latest_file)}")
    print(f"   修改时间: {time.ctime(files_with_time[0][1])}")

    # 加载文件
    wb = openpyxl.load_workbook(latest_file)

    # 检查第一个有数据的工作表
    for sheet_name in wb.sheetnames:
        if sheet_name.isdigit():
            ws = wb[sheet_name]
            print(f"\n📊 工作表: {sheet_name}")

            # 检查R-W列的金额列（V列，第22列）
            amount_col = 22  # V列
            print(f"\n🔍 检查R-W列的金额列（V列）：")

            success_count = 0
            error_count = 0
            data_rows_found = False

            for row in range(2, min(30, ws.max_row + 1)):
                cell = ws.cell(row=row, column=amount_col)
                value = cell.value

                if value is None:
                    continue

                value_str = str(value)
                data_rows_found = True

                # 获取名称列的值来判断行类型
                name_cell = ws.cell(row=row, column=18)  # R列
                name_value = str(name_cell.value) if name_cell.value else ""

                if row == 2:
                    # 表头行
                    if value_str == "金额":
                        print(f"   第{row}行（表头）: {value_str} ✅")
                        success_count += 1
                    else:
                        print(f"   第{row}行（表头）: {value_str} ❌")
                        error_count += 1
                elif "合计" in name_value:
                    # 合计行
                    if value_str.startswith("=SUM("):
                        # 提取SUM范围
                        import re
                        sum_match = re.search(r'SUM\(([^)]+)\)', value_str)
                        if sum_match:
                            sum_range = sum_match.group(1)
                            print(f"   第{row}行（合计）: =SUM({sum_range}) ✅")
                        else:
                            print(f"   第{row}行（合计）: {value_str[:30]}... ✅")
                        success_count += 1
                    else:
                        print(f"   第{row}行（合计）: {value_str} ❌（应该是=SUM公式）")
                        error_count += 1
                elif value_str.startswith("=") and "*" in value_str:
                    # 数据行
                    # 只显示前3个数据行
                    if success_count <= 5:
                        print(f"   第{row}行（数据）: {value_str} ✅")
                    success_count += 1
                elif value_str and not value_str.startswith("="):
                    # 静态值（错误）
                    print(f"   第{row}行: {value_str} ❌（应该是公式）")
                    error_count += 1

            if not data_rows_found:
                print(f"   ⚠️  工作表 {sheet_name} 没有数据")
            else:
                print(f"\n   验证结果: ✅ {success_count} 个正确, ❌ {error_count} 个错误")

                if error_count == 0:
                    print(f"\n   ✅✅✅ 工作表 {sheet_name} 的金额列全部使用公式！")
                else:
                    print(f"\n   ⚠️  工作表 {sheet_name} 发现 {error_count} 个问题")

            # 只检查第一个有数据的工作表
            if data_rows_found:
                break

    print("\n" + "=" * 70)
    print("验证完成")
    print("=" * 70)

    print("\n💡 结论：")
    print("   如果所有行都显示 ✅，说明金额列修复成功！")
    print("   - 表头行显示'金额'")
    print("   - 数据行使用公式 =数量*单价")
    print("   - 合计行使用公式 =SUM(金额范围)")

if __name__ == '__main__':
    import time
    verify_latest_handover()
