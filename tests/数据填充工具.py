"""
数据填充工具 - 自动将送货商文件数据填充到报表
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
import msoffcrypto
import os
import glob
import sys
import argparse
import tempfile
from typing import Optional, Dict, List, Tuple

# 设置输出编码
sys.stdout.reconfigure(encoding="utf-8")


# ============ 样式管理器 ============


class ExcelStyleManager:
    """Excel 样式管理器 - 统一管理所有样式，避免重复创建"""

    def __init__(self):
        self._thin_border = None
        self._header_alignment = None
        self._data_alignment = None
        self._header_font = None

    @property
    def thin_border(self) -> Border:
        """获取细边框样式（懒加载）"""
        if self._thin_border is None:
            self._thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        return self._thin_border

    @property
    def header_alignment(self) -> Alignment:
        """获取表头对齐方式（居中）"""
        if self._header_alignment is None:
            self._header_alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
        return self._header_alignment

    @property
    def data_alignment(self) -> Alignment:
        """获取数据对齐方式（左对齐）"""
        if self._data_alignment is None:
            self._data_alignment = Alignment(
                horizontal="left",
                vertical="center"
            )
        return self._data_alignment

    @property
    def header_font(self) -> Font:
        """获取表头字体（粗体）"""
        if self._header_font is None:
            self._header_font = Font(bold=True)
        return self._header_font


# ============ 数据填充工具 ============


class DataFiller:
    """数据填充工具"""

    # Excel 相关常量
    EXCEL_MAX_COLUMNS = 16384
    HEADER_ROW_OFFSET = 1
    MAX_DAY_SHEET = 31

    # 列相关常量
    SUPPLIER_COLUMN_COUNT = 6  # 名称、单位、数量、单价、总价、备注
    DATA_COLUMN_COUNT = 5  # 不含备注的数据列

    # 要跳过的标题
    SKIP_HEADERS = ["名称", "共计", "送货人：", "收货人："]

    def __init__(self, report_path: str, report_password: Optional[str] = None):
        """
        初始化

        Args:
            report_path: 报表文件路径
            report_password: 报表密码（如果有）
        """
        self.report_path: str = report_path
        self.report_password: Optional[str] = report_password
        self.wb = None
        self.supplier_data: Dict[str, pd.DataFrame] = {}
        self._temp_decrypted_file: Optional[str] = None
        self.styles = ExcelStyleManager()  # ✅ 添加样式管理器

    def load_report(self) -> bool:
        """
        加载报表文件

        Returns:
            bool: 成功返回 True，失败返回 False
        """
        print(f"\n📊 加载报表文件: {self.report_path}")

        # 如果有密码，先尝试解密；如果失败或无密码，直接加载
        if self.report_password:
            try:
                # 使用临时文件
                with tempfile.NamedTemporaryFile(
                    mode="wb", suffix=".xlsx", delete=False
                ) as temp_file:
                    temp_decrypted = temp_file.name

                with open(self.report_path, "rb") as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=self.report_password)
                    with open(temp_decrypted, "wb") as output:
                        office_file.decrypt(output)

                print("✅ 报表解密成功")
                self.report_path = temp_decrypted
                self._temp_decrypted_file = temp_decrypted  # 记录临时文件
            except msoffcrypto.exceptions.DecryptionError as e:
                print(f"⚠️  解密失败（密码错误或文件未加密）: {e}")
                print(f"   尝试直接加载文件...")
                # 解密失败，尝试直接加载（可能是文件没有加密）
                pass
            except Exception as e:
                print(f"⚠️  解密失败: {e}")
                print(f"   尝试直接加载文件...")
                # 其他异常，尝试直接加载
                pass

        # 加载工作簿
        try:
            self.wb = openpyxl.load_workbook(self.report_path)
            print("✅ 成功加载报表")
            print(f"   工作表数量: {len(self.wb.sheetnames)}")
            print(f"   工作表列表: {self.wb.sheetnames}")
            return True
        except OSError as e:
            print(f"❌ 文件读取失败: {e}")
            return False
        except Exception as e:
            print(f"❌ 加载报表失败: {e}")
            return False

    def get_date_sheet_data(self) -> Dict[str, any]:
        """
        从所有数字工作表中获取日期列表

        Returns:
            dict: {
                "dates": ["1", "2", "3", ...],  # 所有数字工作表名称
                "supplierData": {}
            }
        """
        result = {"dates": [], "supplierData": {}}

        if not self.wb:
            print("⚠️ 工作簿未加载")
            return result

        try:
            # 获取所有数字工作表
            numeric_sheets = self.get_all_numeric_sheets()

            if not numeric_sheets:
                print("⚠️ 未找到数字工作表（1-31）")
                return result

            result["dates"] = numeric_sheets
            print(
                f"✅ 获取到 {len(numeric_sheets)} 个日期: {numeric_sheets[:10]}{'...' if len(numeric_sheets) > 10 else ''}"
            )

            return result

        except Exception as e:
            print(f"❌ 获取日期列表失败: {e}")
            import traceback

            traceback.print_exc()
            return result

    def _get_cell_value(self, ws, cell_ref):
        """
        获取单元格的值，支持单元格引用（如 "AE3"）

        Args:
            ws: 工作表对象
            cell_ref: 单元格引用，如 "AE3"

        Returns:
            单元格的值（数字或None）
        """
        try:
            import re
            from openpyxl.utils import column_index_from_string

            # 解析单元格引用，如 "AE3"
            match = re.match(r'([A-Z]+)(\d+)', cell_ref)
            if not match:
                return None

            col_letter = match.group(1)
            row_num = int(match.group(2))
            col_idx = column_index_from_string(col_letter)

            cell = ws.cell(row=row_num, column=col_idx)

            # 如果是公式，尝试获取计算后的值（暂时返回None，因为openpyxl不会自动计算公式）
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                return None

            # 返回数字值
            if cell.value and isinstance(cell.value, (int, float)):
                return cell.value

            return None
        except:
            return None

    def count_supplier_data_in_columns(self, suppliers_list, dates=None):
        """
        统计所有供应商列的合计金额

        Args:
            suppliers_list: 供应商列表，如 [{"name": "万邦蔬菜", "start_column": "AC", "end_column": "AH"}, ...]
            dates: 日期列表（可选，如 ["9", "10", ...]）

        Returns:
            dict: 嵌套字典结构
                  {
                    "AC-AH": {"9": 4100.0, "10": 3200.0, ...},
                    "AI-AN": {"9": 1500.0, ...},
                    ...
                  }
                  如果该日期没有数据，则不包含在字典中
        """
        if not self.wb:
            return {}

        result = {}

        try:
            # 遍历每个供应商
            for supplier in suppliers_list:
                col_start = supplier['start_column']
                col_end = supplier['end_column']
                column_range = f"{col_start}-{col_end}"

                # 调用单列范围的统计方法
                supplier_data = self._count_single_column_range(col_start, col_end, dates)

                if supplier_data:
                    result[column_range] = supplier_data

            return result

        except Exception as e:
            print(f"❌ 检查列数据失败: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def _count_single_column_range(self, col_start: str, col_end: str, dates=None):
        """
        统计单个列范围的合计金额

        Args:
            col_start: 起始列（如"AC"）
            col_end: 结束列（如"AH"）
            dates: 日期列表（可选）

        Returns:
            dict: {"9": 4100.0, "10": 3200.0, ...}
        """
        if not self.wb:
            return {}

        try:
            from openpyxl.utils import column_index_from_string

            # 获取所有数字工作表
            numeric_sheets = self.get_all_numeric_sheets()

            if not numeric_sheets:
                return {}

            # 如果指定了日期，只统计这些日期
            if dates:
                numeric_sheets = [s for s in numeric_sheets if s in dates]

            # 将列字母转换为列索引
            col_start_idx = column_index_from_string(col_start)
            col_end_idx = column_index_from_string(col_end)

            result = {}

            # 遍历每个数字工作表
            for sheet_name in numeric_sheets:
                if sheet_name not in self.wb.sheetnames:
                    continue

                ws = self.wb[sheet_name]

                # 检查是否有数据
                has_data = False
                for row in range(2, ws.max_row + 1):
                    for col_idx in range(col_start_idx, col_end_idx + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None and str(cell.value).strip() != "":
                            # 跳过标题单元格
                            cell_value = str(cell.value).strip()
                            if cell_value not in [
                                "名称",
                                "单位",
                                "数量",
                                "单价",
                                "金额",
                                "备注",
                                "共计",
                                "送货人：",
                                "收货人：",
                            ]:
                                has_data = True
                                break
                    if has_data:
                        break

                if not has_data:
                    continue  # 该日期没有数据，跳过

                # ✅ 逻辑：查找包含"合计"的行，读取其金额列
                total_row = None
                amount_col = col_start_idx + 4  # 第5列（金额列）
                qty_col = col_start_idx + 2    # 第3列（数量列）
                price_col = col_start_idx + 3  # 第4列（单价列）

                # 在供应商列的第一列查找"合计"行（前100行）
                for row in range(1, min(101, ws.max_row + 1)):
                    first_col_cell = ws.cell(row=row, column=col_start_idx)
                    if first_col_cell.value and "合计" in str(first_col_cell.value):
                        total_row = row
                        break

                # 如果找不到"合计"行，使用倒数第三行
                if not total_row:
                    total_row = ws.max_row - 2

                amount_cell = ws.cell(row=total_row, column=amount_col)
                amount_value = None

                # 获取金额值
                if amount_cell.value is not None:
                    # 如果是公式，手动计算求和结果（通过数量×单价）
                    if isinstance(amount_cell.value, str) and amount_cell.value.startswith('='):
                        try:
                            # 简单公式计算：=SUM(AG3:AG51)
                            formula = amount_cell.value
                            if formula.startswith('=SUM('):
                                # 提取范围，如 AG3:AG51
                                range_part = formula[5:-1]  # 去掉 '=SUM(' 和 ')'
                                total = 0

                                if ':' in range_part:
                                    start_ref, end_ref = range_part.split(':')
                                    # 解析起始和结束单元格
                                    import re
                                    start_match = re.match(r'([A-Z]+)(\d+)', start_ref)
                                    end_match = re.match(r'([A-Z]+)(\d+)', end_ref)

                                    if start_match and end_match:
                                        start_row = int(start_match.group(2))
                                        end_row = int(end_match.group(2))

                                        # ✅ 求和：遍历范围内的每一行，通过数量×单价计算金额
                                        for row_idx in range(start_row, end_row + 1):
                                            qty_cell = ws.cell(row=row_idx, column=qty_col)
                                            price_cell = ws.cell(row=row_idx, column=price_col)

                                            # 获取数量和单价
                                            qty = None
                                            price = None

                                            # 获取数量
                                            if qty_cell.value is not None:
                                                if isinstance(qty_cell.value, (int, float)):
                                                    qty = qty_cell.value
                                                else:
                                                    # 检查是否是公式（如 =C3）
                                                    qty_str = str(qty_cell.value).strip()
                                                    if qty_str.startswith('='):
                                                        # 公式，尝试提取数值
                                                        try:
                                                            qty = float(qty_str[1:])  # 去掉 '='
                                                        except:
                                                            pass
                                                    else:
                                                        # 普通字符串，尝试转换
                                                        try:
                                                            qty = float(qty_str)
                                                        except:
                                                            pass

                                            # 获取单价
                                            if price_cell.value is not None:
                                                if isinstance(price_cell.value, (int, float)):
                                                    price = price_cell.value
                                                else:
                                                    # 检查是否是公式
                                                    price_str = str(price_cell.value).strip()
                                                    if price_str.startswith('='):
                                                        # 公式，尝试提取数值
                                                        try:
                                                            price = float(price_str[1:])  # 去掉 '='
                                                        except:
                                                            pass
                                                    else:
                                                        # 普通字符串，尝试转换
                                                        try:
                                                            price = float(price_str)
                                                        except:
                                                            pass

                                            # 计算金额
                                            if qty is not None and price is not None:
                                                total += (qty * price)

                                amount_value = total
                        except Exception:
                            # 公式计算失败
                            pass

                    # 如果还不是数字，尝试直接转换
                    if amount_value is None:
                        try:
                            amount_value = float(amount_cell.value)
                        except (ValueError, TypeError):
                            amount_value = None

                # 只有有金额时才添加到结果
                if amount_value is not None and amount_value > 0:
                    result[sheet_name] = amount_value

            return result

        except Exception as e:
            print(f"❌ 检查列数据失败: {e}")
            import traceback

            traceback.print_exc()
            return {}

    def check_existing_data(self, dates, col_start, col_end):
        """
        检查指定列范围在指定日期是否已有数据（优化版）

        Args:
            dates: 日期列表，如 ["9", "10", "11", "12"]
            col_start: 起始列，如 "AC"
            col_end: 结束列，如 "AH"

        Returns:
            dict: {"9": True/False, "10": True/False, ...}
        """
        print(f"\n🔍 [DEBUG] check_existing_data() 被调用")
        print(f"   日期列表: {dates}")
        print(f"   列范围: {col_start}-{col_end}")
        print(f"   工作表列表: {self.wb.sheetnames if self.wb else 'None'}")

        if not self.wb:
            print(f"   ❌ workbook 未初始化")
            return {date: False for date in dates}

        try:
            from openpyxl.utils import column_index_from_string

            col_start_idx = column_index_from_string(col_start)
            col_end_idx = column_index_from_string(col_end)
            print(f"   列索引: {col_start_idx}-{col_end_idx}")

            result = {}

            for date in dates:
                print(f"\n   检查日期 '{date}':")
                if date not in self.wb.sheetnames:
                    print(f"     ❌ 工作表 '{date}' 不存在")
                    result[date] = False
                    continue

                ws = self.wb[date]
                print(f"     ✅ 工作表 '{date}' 存在，最大行数: {ws.max_row}")
                has_data = False
                found_values = []

                # ✅ 性能优化：使用 iter_rows 批量读取单元格
                # 一次性读取所有目标列的数据，减少 API 调用次数
                for row in ws.iter_rows(
                    min_row=2,  # 从第2行开始
                    max_row=ws.max_row,
                    min_col=col_start_idx,
                    max_col=col_end_idx
                ):
                    for cell in row:
                        if cell.value is not None and str(cell.value).strip() != "":
                            cell_value = str(cell.value).strip()
                            print(
                                f"     找到数据: 行{row}, 列{col_idx}, 值='{cell_value}'"
                            )
                            found_values.append(cell_value)
                            # 使用常量 SKIP_HEADERS
                            if cell_value not in self.SKIP_HEADERS:
                                has_data = True
                                print(f"     ✅ 该值不是表头，标记为有数据")
                                break
                    if has_data:
                        break

                if not has_data and found_values:
                    print(
                        f"     ⚠️  找到 {len(found_values)} 个值，但都是表头值: {found_values[:5]}"
                    )

                result[date] = has_data
                print(f"     结果: {'有数据' if has_data else '无数据'}")

            print(f"\n🔍 [DEBUG] 最终结果: {result}")
            return result

        except Exception as e:
            print(f"❌ 检查现有数据失败: {e}")
            import traceback

            traceback.print_exc()
            return {date: False for date in dates}

    def is_data_cell(
        self, cell_value: any, skip_headers: Optional[List[str]] = None
    ) -> bool:
        """
        判断单元格是否为有效数据

        Args:
            cell_value: 单元格值
            skip_headers: 要跳过的标题列表

        Returns:
            bool: 是否为有效数据
        """
        if skip_headers is None:
            skip_headers = self.SKIP_HEADERS

        if cell_value is None or str(cell_value).strip() == "":
            return False

        cell_str = str(cell_value).strip()
        return cell_str not in skip_headers

    def find_data_rows(
        self,
        ws,
        start_row: int,
        end_row: int,
        col_start: int,
        col_end: int,
    ) -> List[int]:
        """
        查找指定范围内有数据的行

        Args:
            ws: 工作表
            start_row: 起始行
            end_row: 结束行
            col_start: 起始列
            col_end: 结束列

        Returns:
            list: 有数据的行号列表
        """
        data_rows = []
        for row in range(start_row, end_row + 1):
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row, column=col)
                if self.is_data_cell(cell.value):
                    data_rows.append(row)
                    break
        return data_rows

    def get_all_numeric_sheets(self) -> List[str]:
        """
        获取所有名称为数字的工作表，按数字排序

        Returns:
            list: ["1", "2", "3", ..., "31"]
        """
        if not self.wb:
            return []

        numeric_sheets = []
        for sheet_name in self.wb.sheetnames:
            if sheet_name.isdigit():
                numeric_sheets.append((int(sheet_name), sheet_name))

        numeric_sheets.sort()  # 按数字排序
        result = [name for num, name in numeric_sheets]

        if result:
            print(
                f"   📋 找到 {len(result)} 个数字工作表: {result[:10]}{'...' if len(result) > 10 else ''}"
            )

        return result

    def safe_get_value(self, row_data, key, default=""):
        """
        安全获取值，处理 None 和 NaN

        Args:
            row_data: 行数据（dict或Series）
            key: 键名
            default: 默认值

        Returns:
            处理后的值
        """
        value = row_data.get(key, default)
        if pd.isna(value) or value is None:
            return default
        return str(value).strip()

    def normalize_date_str(self, date_str: Optional[str]) -> Optional[str]:
        """
        标准化日期字符串（修复冲突问题）

        Args:
            date_str: 原始日期字符串（如"9", "10.9", "10.10"）

        Returns:
            标准化后的日期字符串（如"9", "9", "10"），如果无效则返回 None
        """
        import re

        # 检查输入是否为 None
        if date_str is None:
            print(f"  ⚠️  日期为 None")
            return None

        try:
            # 转换为字符串
            date_str_str = str(date_str)

            # 如果包含小数点，提取小数点后的部分
            if "." in date_str_str:
                parts = date_str_str.split(".")
                if len(parts) == 2:
                    result = parts[1]
                else:
                    result = date_str_str
            else:
                result = date_str_str

            # 验证是否为有效的日期（1-31）
            if result.isdigit():
                day = int(result)
                if 1 <= day <= 31:
                    return result
                else:
                    print(f"  ⚠️  无效日期: {date_str} -> {result} (超出1-31范围)")
                    return None
            else:
                print(f"  ⚠️  无效日期格式: {date_str}")
                return None

        except Exception as e:
            print(f"  ⚠️  日期标准化失败: {date_str} - {e}")
            return None

    def read_supplier_file(self, supplier_file_path: str) -> None:
        """
        读取送货商文件（无任何限制条件）

        从A1到F列最后一行，读取所有数据，不加任何判断条件

        Args:
            supplier_file_path: 送货商文件路径
        """
        print(f"\n📂 读取送货商文件: {os.path.basename(supplier_file_path)}")

        try:
            # 读取所有工作表
            xls = pd.ExcelFile(supplier_file_path)

            for sheet_name in xls.sheet_names:
                # ✅ 修改1：不跳过第1行，从头开始读取（不把第1行当作列名）
                df = pd.read_excel(
                    supplier_file_path, sheet_name=sheet_name, header=None
                )

                # ✅ 修改2：不删除空行
                # df = df.dropna(how='all')  # 删除这行

                # ✅ 修改3：不检查列数
                # if len(df.columns) < 5:  # 删除这些行
                #     print(f"  ⚠️ 工作表 '{sheet_name}': 列数不足，跳过")
                #     continue

                # ✅ 修改4：不删除最后一行
                # df = df.iloc[:-1]  # 删除这行

                # ✅ 修改5：取前6列（A-F列），而不是5列
                if len(df.columns) >= 6:
                    df = df.iloc[:, :6]  # A-F共6列
                else:
                    # 如果列数不足6列，取所有列
                    df = df.iloc[:, : len(df.columns)]

                # 设置列名（改为6列）
                column_names = ["名称", "单位", "数量", "单价", "总价", "备注"]
                df.columns = column_names[: len(df.columns)]

                # ✅ 修改6：删除完全空的行（所有列都是空的行）
                # 但保留部分有数据的行
                df = df.dropna(how='all')  # 删除所有列都是NaN的行
                # 重置索引
                df = df.reset_index(drop=True)

                # 直接存储数据
                # 始终存储，即使为空
                # 标准化工作表名作为日期
                normalized_date = self.normalize_date_str(sheet_name)

                # 跳过无效日期（None值）
                if normalized_date is None:
                    print(f"  ⚠️  工作表 '{sheet_name}' 日期无效，跳过")
                    continue

                # 如果该日期已存在数据，追加数据
                if normalized_date in self.supplier_data:
                    self.supplier_data[normalized_date] = pd.concat(
                        [self.supplier_data[normalized_date], df], ignore_index=True
                    )
                    print(
                        f"  ✅ 工作表 '{sheet_name}' -> 日期 '{normalized_date}': {len(df)} 条记录（追加）"
                    )
                else:
                    self.supplier_data[normalized_date] = df
                    print(
                        f"  ✅ 工作表 '{sheet_name}' -> 日期 '{normalized_date}': {len(df)} 条记录"
                    )

        except Exception as e:
            print(f"  ❌ 读取失败: {e}")
            import traceback

            traceback.print_exc()

    def find_last_data_row(
        self, ws, start_row: int, col_start: int, col_end: int
    ) -> int:
        """
        找到指定列范围内最后一个有数据的行

        Args:
            ws: 工作表对象
            start_row: 开始行
            col_start: 开始列（1-based）
            col_end: 结束列（1-based）

        Returns:
            最后一个有数据的行号
        """
        last_row = start_row - 1

        for row in range(start_row, ws.max_row + 1):
            # 检查该行在指定列范围内是否有数据
            has_data = False
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    # 跳过"共计"和"名称"这样的标题行
                    if cell.value not in ["共计", "名称", "送货人：", "收货人："]:
                        has_data = True
                        break

            if has_data:
                last_row = row

        return last_row

    def fill_data_to_report(self, date_str: str, col_start: int, col_end: int) -> bool:
        """
        将送货商数据填充到报表（带格式）

        Args:
            date_str: 日期字符串（如"9", "10"）
            col_start: 起始列号（1-based，如17代表Q列）
            col_end: 终止列号（1-based，如21代表U列）
        """
        # 检查报表中是否有对应的工作表
        if date_str not in self.wb.sheetnames:
            print(f"  ⚠️ 报表中没有工作表 '{date_str}'，跳过")
            return False

        # 检查是否有对应日期的送货商数据
        if date_str not in self.supplier_data:
            return False

        ws = self.wb[date_str]
        df_supplier = self.supplier_data[date_str]

        print(f"  📝 填充到工作表 '{date_str}'")

        # 【改进】检查目标列是否已有数据
        has_existing_data = False
        for row in range(2, min(10, ws.max_row + 1)):  # 检查前10行
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    has_existing_data = True
                    break
            if has_existing_data:
                break

        if has_existing_data:
            # 已有数据，追加到末尾
            print(f"     检测到已有数据，将追加到末尾")
            last_data_row = self.find_last_data_row(ws, 2, col_start, col_end)
            print(f"     最后一行数据：第{last_data_row}行")

            # 直接在最后一行数据后面插入表头
            header_row = last_data_row + 1
            ws.insert_rows(header_row, 1)

            # 追加模式下不需要空白行
            blank_row = None
        else:
            # 没有数据，首次填充（需要清空可能存在的合并单元格）
            print(f"     目标列为空，首次填充")
            print(f"     清空目标列（第{col_start}列到第{col_end}列）的旧数据...")

            # 获取所有合并的单元格范围
            merged_cells_ranges = list(ws.merged_cells.ranges)
            print(f"     检测到 {len(merged_cells_ranges)} 个合并单元格区域")

            # 记录需要取消合并的范围（只处理与目标列重叠的合并区域）
            merges_to_unmerge = []

            for merged_range in merged_cells_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                # 检查合并范围是否与目标列重叠
                if not (max_col < col_start or min_col > col_end):
                    # 记录这个合并范围
                    merges_to_unmerge.append(str(merged_range))

            # 取消所有与目标列重叠的合并单元格
            if merges_to_unmerge:
                print(f"     取消 {len(merges_to_unmerge)} 个合并单元格...")
                for merge_range_str in merges_to_unmerge:
                    try:
                        ws.unmerge_cells(merge_range_str)
                    except Exception as e:
                        print(f"     警告：无法取消合并 {merge_range_str}: {e}")

            # 清空所有数据（现在没有合并单元格了）
            cleared_count = 0
            for row in range(2, ws.max_row + 1):
                for col in range(col_start, col_end + 1):
                    try:
                        ws.cell(row=row, column=col).value = None
                        cleared_count += 1
                    except Exception as e:
                        print(
                            f"     警告：无法清空单元格 {ws.cell(row=row, column=col).coordinate}: {e}"
                        )

            print(f"     清空完成：{cleared_count} 个单元格")

            # 先插入一个空白行作为间隔
            blank_row = 2
            ws.insert_rows(blank_row, 1)

            # 再插入表头行
            header_row = blank_row + 1
            ws.insert_rows(header_row, 1)

        # 计算列数（数据列数）
        num_data_cols = 5  # 名称、单位、数量、单价、总价
        num_cols = min(col_end - col_start + 1, num_data_cols)

        # 填充表头
        headers = ["名称", "单位", "数量", "单价", "总价"]
        for i in range(num_cols):
            ws.cell(row=header_row, column=col_start + i).value = headers[i]

        # 应用表头样式（加粗 + 边框）
        from openpyxl.styles import Font, Border, Side, Alignment

        # ✅ 使用样式管理器
        for col in range(col_start, col_start + num_cols):
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.styles.header_font
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.header_alignment

        # 从表头的下一行开始填充数据
        start_row = header_row + 1

        if blank_row is not None:
            print(
                f"     在第{blank_row}行插入空白行，从第{start_row}行开始填充（共{len(df_supplier)}条记录）"
            )
        else:
            print(f"     从第{start_row}行开始追加填充（共{len(df_supplier)}条记录）")

        # 填充数据（带格式）
        for idx, (_, row) in enumerate(df_supplier.iterrows()):
            row_num = start_row + idx

            # 填充数据（根据列数动态填充）
            data_values = [
                row["名称"],
                row["单位"],
                row["数量"],
                row["单价"],
                row["总价"],
            ]
            for i in range(num_cols):
                ws.cell(row=row_num, column=col_start + i).value = data_values[i]

            # ✅ 应用边框样式（使用样式管理器）
            for col in range(col_start, col_start + num_cols):
                cell = ws.cell(row=row_num, column=col)
                cell.border = self.styles.thin_border
                cell.alignment = self.styles.data_alignment

        print(f"     ✅ 成功填充 {len(df_supplier)} 条记录（含格式）")
        return True

    def read_handover_file(self, handover_file_path: str):
        """
        读取交接单文件（从第2行开始到倒数第2行结束）

        Args:
            handover_file_path: 交接单文件路径
        """
        print(f"\n📂 读取交接单文件: {os.path.basename(handover_file_path)}")

        # 清空之前的供应商数据
        self.supplier_data = {}

        try:
            # 读取所有工作表
            xls = pd.ExcelFile(handover_file_path)

            for sheet_name in xls.sheet_names:
                # ✅ 读取工作表：从第2行开始读取（skiprows=1）
                df = pd.read_excel(
                    handover_file_path, sheet_name=sheet_name, skiprows=1
                )

                # ✅ 移除：删除空行
                # df = df.dropna(how='all')  # 删除这行

                # ✅ 移除：检查列数
                # if len(df.columns) < 5:  # 删除这些行
                #     print(f"  ⚠️ 工作表 '{sheet_name}': 列数不足，跳过")
                #     continue

                # ✅ 保留：截取到倒数第2行（删除最后一行）
                df = df.iloc[:-1]

                # ✅ 移除：限制列数（取所有列，不限制为5列）
                # df = df.iloc[:, :5]  # 删除这行

                # 设置列名（支持任意列数）
                if len(df.columns) >= 5:
                    column_names = ["名称", "单位", "数量", "单价", "总价"]
                    if len(df.columns) >= 6:
                        column_names.append("备注")
                    df.columns = column_names[: len(df.columns)]
                else:
                    # 列数不足5列，使用默认列名
                    df.columns = [f"列{i+1}" for i in range(len(df.columns))]

                # ✅ 移除：清理数据（不过滤空行）
                # df = df[df['名称'].notna()]  # 删除这行
                # df = df[df['名称'] != '']  # 删除这行

                # ✅ 移除：检查数据是否为空
                # if len(df) > 0:  # 删除这个判断，始终存储

                # 标准化工作表名作为日期
                normalized_date = self.normalize_date_str(sheet_name)

                # 如果该日期已存在数据，追加数据
                if normalized_date in self.supplier_data:
                    self.supplier_data[normalized_date] = pd.concat(
                        [self.supplier_data[normalized_date], df], ignore_index=True
                    )
                    print(
                        f"  ✅ 工作表 '{sheet_name}' -> 日期 '{normalized_date}': {len(df)} 条记录（追加）"
                    )
                else:
                    self.supplier_data[normalized_date] = df
                    print(
                        f"  ✅ 工作表 '{sheet_name}' -> 日期 '{normalized_date}': {len(df)} 条记录"
                    )

        except Exception as e:
            print(f"  ❌ 读取失败: {e}")
            import traceback

            traceback.print_exc()

        return list(self.supplier_data.keys())  # 返回所有有数据的日期

    def column_letter_to_number(self, col_letter: str) -> int:
        """
        将列字母转换为列号

        Args:
            col_letter: 列字母（如 'A', 'Q', 'V'）

        Returns:
            列号（1-based，A=1, Q=17）
        """
        result = 0
        for char in col_letter.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result

    def fill_all(self, start_column="Q", end_column="U"):
        """
        填充所有数据

        Args:
            start_column: 起始列字母（如 'Q'）
            end_column: 终止列字母（如 'U'）
        """
        print("\n" + "=" * 70)
        print("🚀 开始填充数据")
        print("=" * 70)

        if not self.supplier_data:
            print("⚠️ 没有送货商数据")
            return

        # 转换列字母为列号
        col_start = self.column_letter_to_number(start_column)
        col_end = self.column_letter_to_number(end_column)

        print(
            f"\n📋 目标列范围: {start_column}(列{col_start}) 到 {end_column}(列{col_end})"
        )

        # 过滤掉 None 键并排序
        valid_dates = [d for d in self.supplier_data.keys() if d is not None]
        valid_dates_sorted = sorted(valid_dates, key=lambda x: (x is None, x))

        print(f"\n📋 需要填充的日期: {valid_dates_sorted}")

        success_count = 0
        for date_str in valid_dates_sorted:
            if self.fill_data_to_report(date_str, col_start, col_end):
                success_count += 1

        print("\n" + "=" * 70)
        print(f"✅ 填充完成！成功处理 {success_count}/{len(valid_dates)} 个日期")
        print("=" * 70)

    def fill_data_smart(
        self, column_range, dates, col_start, col_end, mode="overwrite"
    ):
        """
        统一的智能填充数据

        Args:
            column_range: 列范围标识（用于日志，格式："AC-AH"）
            dates: 送货商文件中包含的日期列表（如 ["9", "10", "11"]）
            col_start: 起始列字母（如"AC"）
            col_end: 结束列字母（如"AH"）
            mode: 保留参数以兼容旧代码（不再使用）

        核心逻辑：
            1. 清空目标列在所有1-31日工作表的数据
            2. 只在送货商文件包含的日期填充数据
        """
        print(f"\n{'='*70}")
        print(f"🚀 开始填充：{column_range}")
        print(f"   检测到的日期: {dates}")
        print(f"   列范围: {col_start}-{col_end}")
        print(f"{'='*70}")

        # 1. 清空目标列在所有数字工作表的数据
        self._clear_all_data_in_column_range(col_start, col_end)

        # 转换列字母为列号（用于填充方法）
        from openpyxl.utils import column_index_from_string

        col_start_idx = column_index_from_string(col_start)
        col_end_idx = column_index_from_string(col_end)

        # 2. 只在送货商文件包含的日期填充数据
        success_count = 0
        for date in dates:
            if date not in self.supplier_data:
                print(f"  ⚠️  日期 {date} 没有对应的数据，跳过")
                continue

            if date not in self.wb.sheetnames:
                print(f"  ⚠️  报表中没有工作表 '{date}'，跳过")
                continue

            ws = self.wb[date]
            df_supplier = self.supplier_data[date]

            # 使用现有的填充方法（传递列号而不是列字母）
            self._fill_data_to_worksheet(ws, df_supplier, col_start_idx, col_end_idx)
            success_count += 1
            print(f"  ✅ 已填充日期 {date} 的数据")

        print(f"\n{'='*70}")
        print(f"✅ 填充完成！成功处理 {success_count}/{len(dates)} 个日期")
        print(f"{'='*70}")

    def _clear_all_data_in_column_range(self, col_start, col_end):
        """
        清空目标列在所有数字工作表中的数据

        Args:
            col_start: 起始列字母（如"AC"）
            col_end: 结束列字母（如"AH"）
        """
        from openpyxl.utils import column_index_from_string

        col_start_idx = column_index_from_string(col_start)
        col_end_idx = column_index_from_string(col_end)

        print(
            f"   🔍 清空列范围: {col_start}(索引{col_start_idx}) 到 {col_end}(索引{col_end_idx})"
        )

        # 获取所有数字工作表
        numeric_sheets = self.get_all_numeric_sheets()

        # 遍历所有数字工作表（1-31）
        for sheet_name in numeric_sheets:
            if sheet_name not in self.wb.sheetnames:
                continue

            ws = self.wb[sheet_name]

            # 先处理合并单元格：取消所有与目标列范围重叠的合并单元格（包括第1行）
            merged_cells_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_cells_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                # 检查合并区域是否与目标列范围重叠
                if not (max_col < col_start_idx or min_col > col_end_idx):
                    # 有重叠，取消合并
                    try:
                        ws.unmerge_cells(str(merged_range))
                    except Exception:
                        # 如果无法取消合并，跳过
                        pass

            # 清空该工作表中目标列的数据（从第1行开始）
            for row in range(1, ws.max_row + 1):
                for col_idx in range(col_start_idx, col_end_idx + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value is not None:
                            cell.value = None
                    except Exception:
                        # 某些特殊单元格可能无法修改，跳过
                        pass

        # 输出固定提示信息（不显示统计数字）
        print("已清空原有数据，将添加上传文件数据。")

    def _fill_overwrite(self, ws, date, df_supplier, col_start_idx, col_end_idx):
        """
        覆盖模式实现：清空目标列后填充新数据

        ⚠️ 已废弃：该方法已被统一逻辑替代
        新逻辑：在 fill_data_smart() 中统一清空所有数据，然后只填充指定的日期
        保留此方法仅为兼容性
        """
        # 覆盖模式：清空后填充
        print(f"  📝 工作表 '{date}': 覆盖模式")

        # 1. 清空目标列的旧数据
        print(f"     清空列范围 {col_start_idx}-{col_end_idx} 的旧数据...")
        cleared_count = 0
        for row in range(2, ws.max_row + 1):
            for col_idx in range(col_start_idx, col_end_idx + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.value = None
                    cleared_count += 1

        if cleared_count > 0:
            print(f"     清空了 {cleared_count} 个单元格")

        # 2. 使用现有的 fill_data_to_report 逻辑填充数据
        self._fill_data_to_worksheet(ws, df_supplier, col_start_idx, col_end_idx)

    def _fill_append(self, ws, date, df_supplier, col_start_idx, col_end_idx):
        """
        追加模式实现：只填充空白单元格，保留已有数据

        ⚠️ 已废弃：该方法已被统一逻辑替代
        新逻辑：在 fill_data_smart() 中统一清空所有数据，然后只填充指定的日期
        保留此方法仅为兼容性
        """
        print(f"  📝 工作表 '{date}': 追加模式")

        # 找到第一个空白行或最后一行数据之后
        last_data_row = 1
        for row in range(2, ws.max_row + 1):
            row_has_data = False
            for col_idx in range(col_start_idx, col_end_idx + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and str(cell.value).strip() != "":
                    if str(cell.value).strip() not in [
                        "名称",
                        "共计",
                        "送货人：",
                        "收货人：",
                    ]:
                        row_has_data = True
                        break
            if row_has_data:
                last_data_row = row

        # 从最后一行数据之后开始填充
        start_row = last_data_row + 1

        # 检查是否需要插入表头
        needs_header = True
        for row in range(2, min(start_row, ws.max_row + 1)):
            header_cell = ws.cell(row=row, column=col_start_idx)
            if header_cell.value == "名称":
                needs_header = False
                break

        if needs_header and start_row <= ws.max_row:
            # 插入表头
            print(f"     在第 {start_row} 行插入表头")
            ws.insert_rows(start_row, 1)

            # 填充表头
            headers = ["名称", "单位", "数量", "单价", "总价"]
            num_cols = min(col_end_idx - col_start_idx + 1, 5)
            for i in range(num_cols):
                ws.cell(row=start_row, column=col_start_idx + i).value = headers[i]

            # 应用表头样式
            from openpyxl.styles import Font, Border, Side, Alignment

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col_idx in range(col_start_idx, col_start_idx + num_cols):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            start_row += 1

        # 填充数据
        data_start_row = start_row
        print(f"     从第 {data_start_row} 行开始填充 {len(df_supplier)} 条记录")

        from openpyxl.styles import Border, Side, Alignment

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, (_, row_data) in enumerate(df_supplier.iterrows()):
            row_num = data_start_row + idx

            # 填充数据
            data_values = [
                row_data["名称"],
                row_data["单位"],
                row_data["数量"],
                row_data["单价"],
                row_data["总价"],
            ]
            num_cols = min(col_end_idx - col_start_idx + 1, 5)

            for i in range(num_cols):
                ws.cell(row=row_num, column=col_start_idx + i).value = data_values[i]
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        print(f"     ✅ 成功填充 {len(df_supplier)} 条记录")

    def _fill_data_to_worksheet(self, ws, df_supplier, col_start_idx, col_end_idx):
        """
        智能填充数据到工作表，支持公式和特殊行处理

        Args:
            ws: 工作表对象
            df_supplier: 供应商数据的DataFrame
            col_start_idx: 起始列号（1-based）
            col_end_idx: 结束列号（1-based）

        新逻辑：
        - 第1行：保持原样
        - 第2行（表头）：A-D列复制，E列填"金额"，F列复制
        - 中间数据行：A-D列复制，E列填乘法公式，F列复制
        - 倒数第三行（合计）：A-D列复制，E列填求和公式，F列复制
        - 倒数第二行：空行
        - 最后一行：复制所有（送货人、时间）
        """
        from openpyxl.styles import Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter

        # 取消合并单元格
        merged_cells_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_cells_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if not (max_col < col_start_idx or min_col > col_end_idx) and min_row >= 1:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass

        start_row = 1
        total_rows = len(df_supplier)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, (_, row_data) in enumerate(df_supplier.iterrows()):
            row_num = start_row + idx

            # 判断行类型
            if row_num == 1:
                # 第1行：保持原样，复制所有6列
                self._copy_row_all_columns(ws, row_data, col_start_idx, col_end_idx, row_num, thin_border)

            elif row_num == 2:
                # 第2行（表头行）：A-D列复制，E列填"金额"，F列复制
                self._copy_header_row(ws, row_data, col_start_idx, col_end_idx, row_num, thin_border)

            # ✅ 修复：使用索引判断，而不是行号
            elif idx == total_rows - 3:
                # 倒数第三行（合计行）：A-D列复制，E列填求和公式，F列复制
                self._copy_summary_row(ws, row_data, col_start_idx, col_end_idx, row_num, thin_border, start_row)

            # ✅ 修复：移除跳过倒数第2行的逻辑，填充所有数据
            # elif idx == total_rows - 2:
            #     # 倒数第二行：空行（跳过）
            #     pass

            elif idx == total_rows - 1:
                # 最后一行：复制所有列
                self._copy_row_all_columns(ws, row_data, col_start_idx, col_end_idx, row_num, thin_border)

            else:
                # 中间数据行（包括倒数第2行）：A-D列复制，E列填乘法公式，F列复制
                self._copy_data_row_with_formula(ws, row_data, col_start_idx, col_end_idx, row_num, thin_border)

        print(f"     ✅ 成功填充 {len(df_supplier)} 条记录")

    def _copy_row_all_columns(self, ws, row_data, col_start_idx, col_end_idx, row_num, border):
        """复制行所有列（仅复制前4列：名称、单位、数量、单价）"""
        # 只复制前4列
        data_values = [
            self.safe_get_value(row_data, "名称"),
            self.safe_get_value(row_data, "单位"),
            self.safe_get_value(row_data, "数量"),
            self.safe_get_value(row_data, "单价"),
        ]

        # 只填充前4列
        num_cols = min(4, col_end_idx - col_start_idx + 1)

        for i in range(num_cols):
            try:
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                cell.value = data_values[i]
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            except Exception as e:
                print(f"     ⚠️  填充数据时出错 (行{row_num}, 列{i}): {e}")

        # 第5列和第6列清空
        for i in range(4, min(6, col_end_idx - col_start_idx + 1)):
            try:
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                cell.value = ""
                cell.border = border
            except Exception as e:
                pass

    def _copy_header_row(self, ws, row_data, col_start_idx, col_end_idx, row_num, border):
        """复制表头行（第2行）：E列填"金额" """
        # A-D列：复制表头
        header_values = ["名称", "单位", "数量", "单价"]
        for i in range(4):
            try:
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                cell.value = header_values[i]
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            except Exception as e:
                print(f"     ⚠️  填充表头时出错 (行{row_num}, 列{i}): {e}")

        # E列（第5列）：填"金额"
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 4)
            cell.value = "金额"
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception as e:
            print(f"     ⚠️  填充'金额'时出错: {e}")

        # F列（第6列）：填"备注"
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 5)
            cell.value = "备注"
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        except Exception as e:
            print(f"     ⚠️  填充'备注'时出错: {e}")

    def _copy_data_row_with_formula(self, ws, row_data, col_start_idx, col_end_idx, row_num, border):
        """复制数据行：A-D列复制，E列填乘法公式，F列复制"""
        from openpyxl.utils import get_column_letter

        # A-D列：复制原数据（名称、单位、数量、单价）
        data_values = [
            self.safe_get_value(row_data, "名称"),
            self.safe_get_value(row_data, "单位"),
            self.safe_get_value(row_data, "数量"),
            self.safe_get_value(row_data, "单价"),
        ]

        for i in range(4):
            try:
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                cell.value = data_values[i]
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            except Exception as e:
                print(f"     ⚠️  填充数据时出错 (行{row_num}, 列{i}): {e}")

        # E列（第5列）：填乘法公式 =数量×单价
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 4)
            # 获取数量列和单价列的列字母
            qty_col = get_column_letter(col_start_idx + 2)  # 数量列
            price_col = get_column_letter(col_start_idx + 3)  # 单价列
            cell.value = f"={qty_col}{row_num}*{price_col}{row_num}"
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")
        except Exception as e:
            print(f"     ⚠️  填充乘法公式时出错 (行{row_num}): {e}")

        # F列（第6列）：清空（不复制备注，因为备注是我添加的）
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 5)
            cell.value = ""  # 清空备注列
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        except Exception as e:
            print(f"     ⚠️  清空备注列时出错 (行{row_num}): {e}")

    def _copy_summary_row(self, ws, row_data, col_start_idx, col_end_idx, row_num, border, start_row):
        """复制合计行：A-D列复制，E列填求和公式，F列复制"""
        from openpyxl.utils import get_column_letter

        # A-D列：复制原数据
        for i in range(4):
            try:
                cell = ws.cell(row=row_num, column=col_start_idx + i)
                # ✅ 修复：第1列强制填写"合计"
                if i == 0:
                    cell.value = "合计"  # 强制写入"合计"，不检查原数据
                else:
                    cell.value = ""
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            except Exception as e:
                print(f"     ⚠️  填充合计行时出错 (行{row_num}, 列{i}): {e}")

        # E列（第5列）：填求和公式
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 4)
            # 求和范围：从第3行到当前行-1
            sum_col = get_column_letter(col_start_idx + 4)  # E列
            cell.value = f"=SUM({sum_col}{start_row + 2}:{sum_col}{row_num - 1})"
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center")
        except Exception as e:
            print(f"     ⚠️  填充求和公式时出错 (行{row_num}): {e}")

        # F列（第6列）：清空
        try:
            cell = ws.cell(row=row_num, column=col_start_idx + 5)
            cell.value = ""
            cell.border = border
        except Exception as e:
            pass

    def check_supplier_has_data(self, ws, start_col: int, end_col: int) -> bool:
        """
        检查指定列范围内是否有数据（从第2行开始）

        Args:
            ws: 工作表对象
            start_col: 起始列号（1-based）
            end_col: 终止列号（1-based）

        Returns:
            是否有数据
        """
        # 从第2行开始检查（第1行是标题行）
        for row in range(2, ws.max_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    # 跳过非数据行
                    if cell.value not in ["共计", "送货人：", "收货人："]:
                        return True
        return False

    def get_supplier_data_range(self, ws, start_col: int, end_col: int) -> tuple:
        """
        获取供应商数据的实际行范围（从第2行开始到倒数第2行结束）

        Args:
            ws: 工作表对象
            start_col: 起始列号（1-based）
            end_col: 终止列号（1-based）

        Returns:
            (start_row, end_row) 元组，如果没有数据则返回 (None, None)
        """
        start_row = None
        end_row = None

        # 找到第一个有数据的行（从第2行开始）
        for row in range(2, ws.max_row + 1):
            has_data = False
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    has_data = True
                    break
            if has_data:
                start_row = row
                break

        if start_row is None:
            return (None, None)

        # 找到倒数第二个有数据的行（不包含最后一行）
        # 先找到倒数第一个有数据的行
        last_data_row = None
        for row in range(ws.max_row, start_row - 1, -1):
            has_data = False
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip() != "":
                    has_data = True
                    break

            if has_data:
                last_data_row = row
                break

        # 如果找到了最后一行，则end_row设为倒数第二行
        if last_data_row is not None and last_data_row > start_row:
            end_row = last_data_row - 1  # 倒数第二行
        else:
            end_row = start_row  # 只有一行数据

        return (start_row, end_row)

    def _clear_target_columns(self, ws, col_start: int, col_end: int):
        """
        清空目标列的所有数据和格式

        Args:
            ws: 工作表对象
            col_start: 起始列号（1-based）
            col_end: 终止列号（1-based）
        """
        print(f"\n🗑️ 清空目标列（第{col_start}列到第{col_end}列）的旧数据...")

        # 获取所有合并的单元格范围
        merged_cells_ranges = list(ws.merged_cells.ranges)
        print(f"     检测到 {len(merged_cells_ranges)} 个合并单元格区域")

        # 记录需要取消合并的范围（只处理与目标列重叠的合并区域）
        merges_to_unmerge = []

        for merged_range in merged_cells_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            # 检查合并范围是否与目标列重叠
            if not (max_col < col_start or min_col > col_end):
                # 记录这个合并范围
                merges_to_unmerge.append(str(merged_range))

        # 取消所有与目标列重叠的合并单元格
        if merges_to_unmerge:
            print(f"     取消 {len(merges_to_unmerge)} 个合并单元格...")
            for merge_range_str in merges_to_unmerge:
                try:
                    ws.unmerge_cells(merge_range_str)
                except Exception as e:
                    print(f"     警告：无法取消合并 {merge_range_str}: {e}")

        # 清空所有数据（现在没有合并单元格了）
        cleared_count = 0
        for row in range(2, ws.max_row + 1):
            for col in range(col_start, col_end + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.value = None
                    # 清除格式
                    from openpyxl.styles import Font, Border, Alignment

                    cell.font = Font()
                    cell.border = Border()
                    cell.alignment = Alignment()
                    cleared_count += 1
                except Exception as e:
                    print(f"     警告：无法清空单元格 {cell.coordinate}: {e}")

        print(f"     清空完成：{cleared_count} 个单元格")

    def copy_data_to_handover(
        self,
        ws,
        source_start_col: int,
        source_end_col: int,
        target_start_col: int,
        target_end_col: int,
        target_row: int,
    ) -> int:
        """
        将供应商数据复制到Q-V列的指定行

        Args:
            ws: 工作表对象
            source_start_col: 源数据起始列号（1-based）
            source_end_col: 源数据终止列号（1-based）
            target_start_col: 目标起始列号（1-based）
            target_end_col: 目标终止列号（1-based）
            target_row: 目标起始行号（1-based）

        Returns:
            复制的行数
        """
        # 获取源数据范围
        source_start_row, source_end_row = self.get_supplier_data_range(
            ws, source_start_col, source_end_col
        )

        if source_start_row is None or source_end_row is None:
            return 0

        # 计算需要复制的行数
        num_rows = source_end_row - source_start_row + 1

        # 计算源列数和目标列数的最小值
        source_num_cols = source_end_col - source_start_col + 1
        target_num_cols = target_end_col - target_start_col + 1
        num_cols = min(source_num_cols, target_num_cols)

        from openpyxl.styles import Font, Border, Side, Alignment

        # 复制每一行数据
        for row_offset in range(num_rows):
            source_row = source_start_row + row_offset
            target_row_num = target_row + row_offset

            for col_offset in range(num_cols):
                source_col = source_start_col + col_offset
                target_col = target_start_col + col_offset

                source_cell = ws.cell(row=source_row, column=source_col)
                target_cell = ws.cell(row=target_row_num, column=target_col)

                # 复制值
                target_cell.value = source_cell.value

                # 复制格式
                if source_cell.has_style:
                    if source_cell.font:
                        target_cell.font = source_cell.font.copy()
                    if source_cell.border:
                        target_cell.border = source_cell.border.copy()
                    if source_cell.alignment:
                        target_cell.alignment = source_cell.alignment.copy()

        return num_rows

    def compile_handover_from_suppliers(self, suppliers_config):
        """
        从所有供应商列复制数据到R-W列（交接单列）

        Args:
            suppliers_config: 送货商配置

        Returns:
            复制的送货商数量
        """
        # 目标列：R-W（18-23）
        target_start_col = 18  # R列
        target_end_col = 23  # W列

        print("\n" + "=" * 70)
        print("🚀 开始从供应商列复制数据到交接单（R-W列）")
        print("=" * 70)

        if not self.wb.sheetnames:
            print("❌ 没有工作表")
            return 0

        # 处理所有工作表
        total_suppliers_count = 0
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            print(f"\n📋 处理工作表：{sheet_name}")

            # 第一步：清空R-W列（从第2行开始）
            self._clear_target_columns(ws, target_start_col, target_end_col)

            # 第二步：扫描所有送货商，找出有数据的
            suppliers_with_data = []
            for supplier in suppliers_config["suppliers"]:
                start_col = self.column_letter_to_number(supplier["start_column"])
                end_col = self.column_letter_to_number(supplier["end_column"])

                if self.check_supplier_has_data(ws, start_col, end_col):
                    data_range = self.get_supplier_data_range(ws, start_col, end_col)
                    suppliers_with_data.append(
                        {
                            "name": supplier["name"],
                            "start_col": start_col,
                            "end_col": end_col,
                            "data_start_row": data_range[0],
                            "data_end_row": data_range[1],
                        }
                    )
                    print(
                        f"   ✅ 找到数据：{supplier['name']} ({supplier['start_column']}-{supplier['end_column']}列)"
                    )

            print(f"\n找到 {len(suppliers_with_data)} 个有数据的送货商")

            if not suppliers_with_data:
                print("⚠️ 当前工作表没有找到任何送货商数据")
                continue

            # 第三步：按顺序复制数据到R-W列
            target_row = 2  # 从第2行开始（第1行是标题行）
            for idx, supplier_info in enumerate(suppliers_with_data):
                if idx == 0:
                    print(f"\n📋 第1个送货商：{supplier_info['name']}")
                    print(f"   清空R-W列，从第{target_row}行开始粘贴")
                else:
                    print(f"\n📋 第{idx+1}个送货商：{supplier_info['name']}")
                    print(f"   从第{target_row}行开始粘贴（与上一个间隔1行）")

                # 复制数据
                rows_copied = self.copy_data_to_handover(
                    ws,
                    supplier_info["start_col"],
                    supplier_info["end_col"],
                    target_start_col,
                    target_end_col,
                    target_row,
                )

                print(f"   ✅ 复制了 {rows_copied} 行数据")

                # 更新目标行：下一个送货商间隔1个空行
                target_row += rows_copied + 1  # +1 是空行

            total_suppliers_count += len(suppliers_with_data)
            print(
                f"\n✅ 工作表 '{sheet_name}' 处理完成！共复制 {len(suppliers_with_data)} 个送货商的数据"
            )

        print("\n" + "=" * 70)
        print(
            f"✅ 交接单编译完成！所有工作表共复制 {total_suppliers_count} 个送货商的数据"
        )
        print("=" * 70)

        return total_suppliers_count

    def compile_stock_in_from_suppliers(self, suppliers_config):
        """
        从所有供应商列复制数据到入库单列（A-F列）

        Args:
            suppliers_config: 送货商配置

        Returns:
            复制的送货商数量
        """
        # 目标列：A-F（1-6）
        target_start_col = 1  # A列
        target_end_col = 6    # F列

        print("\n" + "=" * 70)
        print("🚀 开始从供应商列复制数据到入库单（A-F列）")
        print("=" * 70)

        if not self.wb.sheetnames:
            print("❌ 没有工作表")
            return 0

        # 处理所有工作表
        total_suppliers_count = 0
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            print(f"\n📋 处理工作表：{sheet_name}")

            # 第一步：清空A-F列（从第2行开始）
            self._clear_target_columns(ws, target_start_col, target_end_col)

            # 第二步：扫描所有送货商，找出有数据的
            suppliers_with_data = []
            for supplier in suppliers_config["suppliers"]:
                start_col = self.column_letter_to_number(supplier["start_column"])
                end_col = self.column_letter_to_number(supplier["end_column"])

                if self.check_supplier_has_data(ws, start_col, end_col):
                    data_range = self.get_supplier_data_range(ws, start_col, end_col)
                    suppliers_with_data.append(
                        {
                            "name": supplier["name"],
                            "start_col": start_col,
                            "end_col": end_col,
                            "data_start_row": data_range[0],
                            "data_end_row": data_range[1],
                        }
                    )
                    print(
                        f"   ✅ 找到数据：{supplier['name']} ({supplier['start_column']}-{supplier['end_column']}列)"
                    )

            print(f"\n找到 {len(suppliers_with_data)} 个有数据的送货商")

            if not suppliers_with_data:
                print("⚠️ 当前工作表没有找到任何送货商数据")
                continue

            # 第三步：按顺序复制数据到A-F列
            target_row = 2  # 从第2行开始（第1行是标题行）
            for idx, supplier_info in enumerate(suppliers_with_data):
                if idx == 0:
                    print(f"\n📋 第1个送货商：{supplier_info['name']}")
                    print(f"   清空A-F列，从第{target_row}行开始粘贴")
                else:
                    print(f"\n📋 第{idx+1}个送货商：{supplier_info['name']}")
                    print(f"   从第{target_row}行开始粘贴（与上一个间隔1行）")

                # 复制数据
                rows_copied = self.copy_data_to_handover(
                    ws,
                    supplier_info["start_col"],
                    supplier_info["end_col"],
                    target_start_col,
                    target_end_col,
                    target_row,
                )

                print(f"   ✅ 复制了 {rows_copied} 行数据")

                # 更新目标行：下一个送货商间隔1个空行
                target_row += rows_copied + 1  # +1 是空行

            total_suppliers_count += len(suppliers_with_data)
            print(
                f"\n✅ 工作表 '{sheet_name}' 处理完成！共复制 {len(suppliers_with_data)} 个送货商的数据"
            )

        print("\n" + "=" * 70)
        print(
            f"✅ 入库单编译完成！所有工作表共复制 {total_suppliers_count} 个送货商的数据"
        )
        print("=" * 70)

        return total_suppliers_count

    def delete_columns_after(self, start_column_letter: str):
        """
        删除指定列之后的所有列的所有内容和格式

        Args:
            start_column_letter: 起始列字母（如 'AB'），删除AB列之后的所有列
        """
        start_col = self.column_letter_to_number(start_column_letter)

        print("\n" + "=" * 70)
        print(f"🗑️ 删除 {start_column_letter} 列之后的所有列")
        print("=" * 70)

        if not self.wb.sheetnames:
            print("❌ 没有工作表")
            return False

        total_deleted = 0

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            if ws.max_column <= start_col:
                print(f"⚠️ 工作表 '{sheet_name}': 没有 {start_column_letter} 列之后的列")
                continue

            print(f"\n📋 处理工作表：{sheet_name}")

            # 计算要删除的列数
            num_cols_to_delete = ws.max_column - start_col
            print(f"   将删除 {start_column_letter} 列之后的 {num_cols_to_delete} 列")

            # 取消合并单元格（先检查是否有合并）
            merged_cells_ranges = list(ws.merged_cells.ranges)
            merges_to_unmerge = []

            for merged_range in merged_cells_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                # 如果合并范围在要删除的列之后，取消合并
                if min_col > start_col:
                    merges_to_unmerge.append(str(merged_range))

            if merges_to_unmerge:
                print(f"   取消 {len(merges_to_unmerge)} 个合并单元格...")
                for merge_range_str in merges_to_unmerge:
                    try:
                        ws.unmerge_cells(merge_range_str)
                    except Exception as e:
                        print(f"   警告：无法取消合并 {merge_range_str}: {e}")

            # 删除列（从右向左删除，避免索引变化）
            deleted_count = 0
            for col in range(ws.max_column, start_col, -1):
                try:
                    ws.delete_cols(col)
                    deleted_count += 1
                except Exception as e:
                    print(f"   警告：无法删除第{col}列: {e}")

            print(f"   ✅ 删除了 {deleted_count} 列")
            total_deleted += deleted_count

        print("\n" + "=" * 70)
        print(f"✅ 删除完成！所有工作表共删除了 {total_deleted} 列")
        print("=" * 70)

        return True

    def save_report(self, output_path: Optional[str] = None) -> Optional[str]:
        """
        保存报表

        Args:
            output_path: 输出路径，如果不指定则覆盖原文件
        """
        if output_path is None:
            # 生成输出文件名
            base_name = os.path.splitext(self.report_path)[0]
            output_path = f"{base_name}_已填充.xlsx"

        try:
            self.wb.save(output_path)
            print(f"\n✅ 报表已保存到: {output_path}")
            return output_path
        except Exception as e:
            print(f"❌ 保存失败: {e}")
            return None

    def cleanup(self):
        """清理临时文件和工作簿资源"""
        # 关闭工作簿
        if self.wb is not None:
            try:
                self.wb.close()
            except Exception:
                pass
            self.wb = None

        # 删除临时解密文件
        if self._temp_decrypted_file and os.path.exists(self._temp_decrypted_file):
            try:
                os.unlink(self._temp_decrypted_file)
                print(f"✅ 已清理临时文件: {self._temp_decrypted_file}")
            except Exception as e:
                print(f"⚠️  清理临时文件失败: {e}")
            finally:
                self._temp_decrypted_file = None

    def __del__(self):
        """析构函数，确保资源释放"""
        self.cleanup()


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description="数据填充工具 - 自动将送货商文件数据填充到报表",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 基本使用
  python 数据填充工具.py --report 报表.xlsx --supplier-folder ./data

  # 带密码的报表
  python 数据填充工具.py --report 报表.xlsx --supplier-folder ./data --password 123456

  # 指定输出文件
  python 数据填充工具.py --report 报表.xlsx --supplier-folder ./data --output 结果.xlsx
        """,
    )

    parser.add_argument("--report", "-r", required=True, help="报表文件路径（必需）")
    parser.add_argument(
        "--supplier-folder", "-s", required=True, help="送货商文件文件夹路径（必需）"
    )
    parser.add_argument(
        "--password", "-p", default="", help="报表密码（可选，默认为空）"
    )
    parser.add_argument(
        "--output", "-o", help='输出文件路径（可选，默认为"报表名_已填充_时间戳.xlsx"）'
    )

    return parser.parse_args()


def main():
    """主函数"""

    # 解析命令行参数
    args = parse_args()

    print("=" * 70)
    print("   数据填充工具")
    print("   自动将送货商数据填充到报表")
    print("=" * 70)

    # 验证输入参数
    if not os.path.exists(args.report):
        print(f"❌ 报表文件不存在: {args.report}")
        return 1

    if not os.path.exists(args.supplier_folder):
        print(f"❌ 送货商文件夹不存在: {args.supplier_folder}")
        return 1

    # 使用命令行参数
    report_path = args.report
    supplier_folder = args.supplier_folder
    report_password = args.password if args.password else None

    # 创建填充工具
    filler = None
    try:
        filler = DataFiller(report_path, report_password)

        # 加载报表
        if not filler.load_report():
            return 1

        # 读取所有送货商文件
        print("\n📂 查找送货商文件...")
        supplier_files = glob.glob(os.path.join(supplier_folder, "*.xls")) + glob.glob(
            os.path.join(supplier_folder, "*.xlsx")
        )

        # 过滤掉报表文件
        supplier_files = [
            f for f in supplier_files if "报表" not in os.path.basename(f)
        ]

        print(f"找到 {len(supplier_files)} 个文件")

        for supplier_file in supplier_files:
            filler.read_supplier_file(supplier_file)

        # 填充所有数据
        filler.fill_all()

        # 保存报表
        if args.output:
            output_path = args.output
        else:
            output_path = os.path.join(
                os.path.dirname(report_path),
                f"金融岛报表_已填充_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            )
        filler.save_report(output_path)

        print("\n" + "=" * 70)
        print("✅ 所有操作完成！")
        print("=" * 70)
        return 0

    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断操作")
        return 130
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback

        traceback.print_exc()
        return 1
    finally:
        # 确保清理资源
        if filler:
            filler.cleanup()


if __name__ == "__main__":
    import sys

    sys.exit(main())
