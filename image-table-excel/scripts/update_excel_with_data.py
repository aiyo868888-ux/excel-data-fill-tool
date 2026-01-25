"""
将提取的数据更新到Excel文件
支持在指定位置插入数据、插入空行、粘贴图片
"""

import argparse
import json
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


class ExcelUpdater:
    """Excel更新器"""

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            print(f"错误: Excel文件不存在: {excel_path}")
            sys.exit(1)

        try:
            self.wb = load_workbook(self.excel_path)
        except Exception as e:
            print(f"错误: 无法打开Excel文件: {e}")
            sys.exit(1)

    def update(
        self,
        data: List[Dict[str, Any]],
        insert_row: int,
        target_sheet: Optional[str] = None,
        insert_blank_row: bool = True,
        image_path: Optional[str] = None,
        image_column: Optional[str] = None
    ):
        """更新Excel数据"""
        # 选择工作表
        ws = self._select_sheet(target_sheet)

        # 检查插入位置
        max_row = ws.max_row
        if insert_row > max_row + 1:
            print(f"警告: 插入行 {insert_row} 超出数据范围，将追加到末尾")
            insert_row = max_row + 1

        # 插入空行（如果需要）
        if insert_blank_row and insert_row <= max_row:
            ws.insert_rows(insert_row)

        # 写入数据
        start_row = insert_row + (1 if insert_blank_row else 0)
        self._write_data(ws, data, start_row)

        # 粘贴图片（如果需要）
        if image_path and image_column:
            self._paste_image(ws, image_path, start_row, image_column)

        # 保存文件
        self.wb.save(self.excel_path)
        print(f"Excel文件已更新: {self.excel_path}")

    def _select_sheet(self, sheet_name: Optional[str]):
        """选择工作表"""
        if sheet_name:
            if sheet_name not in self.wb.sheetnames:
                print(f"警告: 工作表 '{sheet_name}' 不存在，使用第一个工作表")
                sheet_name = self.wb.sheetnames[0]
            return self.wb[sheet_name]
        else:
            return self.wb.active

    def _write_data(self, ws, data: List[Dict[str, Any]], start_row: int):
        """写入数据到工作表"""
        for row_idx, row_data in enumerate(data, start=start_row):
            cells = row_data.get("cells", [])
            for col_idx, cell_value in enumerate(cells, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        print(f"已写入 {len(data)} 行数据到第 {start_row} 行开始的位置")

    def _paste_image(self, ws, image_path: str, row: int, column: str):
        """将图片粘贴到指定单元格"""
        try:
            # 将列字母转换为列索引
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(column)

            # 打开并调整图片大小
            img = PILImage.open(image_path)

            # 调整图片大小以适应单元格
            cell_width = ws.column_dimensions[column].width or 10
            cell_height = ws.row_dimensions[row].height or 15

            # 转换为像素（approximate）
            target_width = int(cell_width * 7)
            target_height = int(cell_height * 1.5)

            img_resized = img.copy()
            img_resized.thumbnail((target_width, target_height), PILImage.Resampling.LANCZOS)

            # 保存临时图片
            temp_img_path = Path(image_path).parent / f"temp_{Path(image_path).name}"
            img_resized.save(temp_img_path)

            # 添加到Excel
            excel_img = XLImage(str(temp_img_path))
            excel_img.anchor = f"{column}{row}"

            ws.add_image(excel_img)

            # 删除临时文件
            temp_img_path.unlink()

            print(f"已将图片粘贴到单元格 {column}{row}")

        except Exception as e:
            print(f"警告: 粘贴图片失败: {e}")


def load_data(json_path: str) -> List[Dict[str, Any]]:
    """从JSON文件加载数据"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"错误: 无法加载数据文件: {e}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='将提取的数据更新到Excel文件')
    parser.add_argument('excel_path', help='Excel文件路径')
    parser.add_argument('--data', required=True, help='提取的数据JSON文件路径')
    parser.add_argument('--insert-row', type=int, default=1, help='插入起始行号（默认: 1）')
    parser.add_argument('--target-sheet', help='目标工作表名称（默认: 活动工作表）')
    parser.add_argument('--no-blank-row', action='store_true', help='不插入空行')
    parser.add_argument('--paste-image', help='粘贴图片到Excel')
    parser.add_argument('--image-column', help='图片粘贴列（如: A, B, C）')

    args = parser.parse_args()

    # 加载数据
    data = load_data(args.data)

    if not data:
        print("错误: 数据文件为空")
        sys.exit(1)

    # 创建更新器
    updater = ExcelUpdater(args.excel_path)

    # 更新Excel
    updater.update(
        data=data,
        insert_row=args.insert_row,
        target_sheet=args.target_sheet,
        insert_blank_row=not args.no_blank_row,
        image_path=args.paste_image,
        image_column=args.image_column
    )


if __name__ == '__main__':
    main()
