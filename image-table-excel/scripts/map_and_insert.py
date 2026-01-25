"""
将图片表格数据映射到中原证券进场材料统计表模板 - 增强版

新功能：
1. 从最下边可用行开始插入
2. 与原有内容间增加空行
3. 只写入表头一致的列
4. 图片只在第一行插入，多图片向后列继续插入
5. 无表格时只插入图片
6. 支持用户自定义内容
"""
import json
import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


# 表头映射：图片列名 -> Excel列索引
HEADER_MAPPING = {
    "序号": 1,      # A列
    "品名": 4,      # D列 - 材料名称
    "规格": 5,      # E列 - 规格型号
    "单位": 6,      # F列
    "数量": 7,      # G列
    "备注": 12      # L列
}

# 列名到列索引的映射（用于用户自定义）
COL_NAME_TO_IDX = {
    "序号": 1,
    "进场时间1": 2,
    "进场时间2": 3,
    "材料名称": 4,
    "规格型号": 5,
    "单位": 6,
    "数量": 7,
    "供货单位": 8,
    "接收人": 9,
    "接收时间": 10,
    "工程部位": 11,
    "备注": 12
}


def find_last_data_row(ws, start_row=5):
    """找到最后一个有数据的行（完整扫描，不遗漏任何数据）"""
    last_data_row = start_row - 1
    
    # 获取工作表的实际最大行（包括所有使用过的行）
    max_row = ws.max_row
    
    print(f"  [调试] 工作表 max_row: {max_row}")
    
    # 向下查找所有行，找到最后一个有数据的行（不因空行中断）
    # 扫描范围：从 start_row 到 max_row，确保不遗漏任何数据
    for row in range(start_row, max_row + 1):
        # 检查整行是否为空（检查前13列：A-M）
        row_has_data = False
        for col in range(1, 14):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip() != "":
                row_has_data = True
                break
        
        if row_has_data:
            last_data_row = row
            print(f"  [调试] 第{row}行有数据")
    
    print(f"  [调试] 最终确定最后数据行: {last_data_row}")
    return last_data_row


def is_valid_table_row(row_data):
    """判断是否为有效的表格数据行"""
    cells = row_data.get("cells", [])

    # 至少需要4列（序号、品名、规格、单位）
    if len(cells) < 4:
        return False

    # 第一列应该是数字（序号）
    first_col = str(cells[0]).strip()
    if not first_col.isdigit():
        return False

    # 排除表头行
    header_keywords = ["序号", "品名", "规格", "单位", "数量", "备注"]
    if any(keyword in str(cells[0]) or keyword in str(cells[1])
           for keyword in header_keywords):
        return False

    # 排除汇总行
    summary_keywords = ["合计", "零元整", "总计", "元", "备注："]
    row_text = " ".join(str(c) for c in cells)
    if any(keyword in row_text for keyword in summary_keywords):
        return False

    return True


def read_header_mapping(ws, header_row=4):
    """读取Excel表头，返回列索引到列名的映射"""
    header_mapping = {}  # 列索引 -> 列名
    for col in range(1, 14):  # A到M列
        cell = ws.cell(row=header_row, column=col)
        if cell.value:
            header_mapping[col] = str(cell.value).strip()
    return header_mapping


def map_row_to_columns(row_data, header_mapping):
    """按照Excel表头顺序映射数据"""
    # 先过滤无效行
    if not is_valid_table_row(row_data):
        return {}

    mapped = {}
    cells = row_data.get("cells", [])

    # 按照表头顺序填充数据（直接按顺序插入）
    for col_idx, col_name in header_mapping.items():
        cell_idx = col_idx - 1  # 列索引从0开始
        if cell_idx < len(cells) and cells[cell_idx]:
            mapped[col_idx] = cells[cell_idx]

    return mapped


def insert_images_to_row(ws, row_num, image_paths, start_col=13):
    """在指定行插入多张图片（原图），从start_col开始向后排列"""
    for idx, image_path in enumerate(image_paths):
        col_idx = start_col + idx
        col_letter = get_column_letter(col_idx)

        try:
            # 直接使用原图，不压缩
            excel_img = XLImage(str(image_path))
            excel_img.anchor = f"{col_letter}{row_num}"
            ws.add_image(excel_img)
            print(f"  已插入原图: {Path(image_path).name} -> {col_letter}{row_num}")
        except Exception as e:
            print(f"  警告: 插入图片 {image_path} 失败: {e}")

    # 设置行高以适应原图（较大高度）
    ws.row_dimensions[row_num].height = 80


def insert_image_to_cell(ws, row_num, col_idx, image_path):
    """在指定单元格插入单张图片（原图）"""
    col_letter = get_column_letter(col_idx)

    try:
        # 直接使用原图，不压缩
        excel_img = XLImage(str(image_path))
        excel_img.anchor = f"{col_letter}{row_num}"
        ws.add_image(excel_img)
    except Exception as e:
        print(f"  警告: 插入图片 {image_path} 到 {col_letter}{row_num} 失败: {e}")

    # 设置行高以适应原图
    if ws.row_dimensions[row_num].height is None or ws.row_dimensions[row_num].height < 80:
        ws.row_dimensions[row_num].height = 80


def has_table_content(image_data):
    """检查图片数据是否包含表格内容"""
    if not image_data or len(image_data) == 0:
        return False

    # 检查是否有有效的数据行（至少2列有内容）
    for row in image_data:
        cells = row.get("cells", [])
        non_empty = [c for c in cells if c and c.strip()]
        if len(non_empty) >= 2:
            return True

    return False


def apply_user_custom_data(row_data, user_custom_data):
    """应用用户自定义的数据"""
    for col_name, value in user_custom_data.items():
        if col_name in COL_NAME_TO_IDX:
            col_idx = COL_NAME_TO_IDX[col_name]
            row_data[col_idx] = value

    return row_data


def process_and_insert(
    excel_path: str,
    image_paths: list = None,
    data_file: str = None,
    user_custom: dict = None
):
    """主函数：处理图片和数据并插入Excel"""

    # 1. 加载表格数据（如果有）
    image_data = []
    if data_file and Path(data_file).exists():
        print(f"读取数据文件: {data_file}")
        with open(data_file, 'r', encoding='utf-8') as f:
            image_data = json.load(f)
        print(f"  共 {len(image_data)} 行数据")

    # 2. 检查是否有表格内容
    has_table = has_table_content(image_data)
    if has_table:
        print("  检测到表格内容")
    else:
        print("  未检测到表格内容，将只插入图片")

    # 3. 打开Excel
    print(f"打开Excel文件: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 4. 读取表头映射
    header_mapping = read_header_mapping(ws)
    print(f"  表头列数: {len(header_mapping)}")
    print(f"  表头内容: {list(header_mapping.values())}")

    # 5. 找到最后一个数据行
    last_data_row = find_last_data_row(ws)
    print(f"  最后一个数据行: {last_data_row}")

    # 6. 【修改】不使用 insert_rows，改为直接追加
    # 在最后数据行后留一个空行作为分隔，然后开始写入
    insert_row = last_data_row + 2  # +1为空行，+2为新数据起始行
    print(f"  追加位置: 第{insert_row}行（第{insert_row-1}行为空行分隔，历史内容不受影响）")

    # 7. 如果有表格数据，写入数据
    if has_table:
        print("\n写入表格数据:")
        valid_row_count = 0
        rows_with_images = []  # 记录需要插入图片的行

        for idx, row_data in enumerate(image_data):
            # 映射数据到列（传入header_mapping）
            mapped = map_row_to_columns(row_data, header_mapping)

            # 跳过无效行（返回空字典）
            if not mapped:
                continue

            valid_row_count += 1
            current_row = insert_row + valid_row_count - 1

            # 应用用户自定义数据
            if user_custom:
                mapped = apply_user_custom_data(mapped, user_custom)

            # 写入单元格
            for col_idx, value in mapped.items():
                if value:  # 只写入非空值
                    col_letter = get_column_letter(col_idx)
                    ws.cell(row=current_row, column=col_idx, value=value)

            # 记录需要插入图片的行
            rows_with_images.append(current_row)

            # 显示写入的数据
            cells_display = [str(mapped.get(i, "")) for i in sorted(mapped.keys())]
            print(f"  行{current_row}: {', '.join(cells_display)}")

        print(f"  共过滤并写入 {valid_row_count} 行有效数据")

    # 8. 插入图片到对应的数据行（M列）
    if image_paths:
        print(f"\n插入图片到表格内:")
        # 为每个有图片的行插入图片
        for row_num in rows_with_images:
            for img_idx, img_path in enumerate(image_paths):
                insert_image_to_cell(ws, row_num, 13, img_path)
                print(f"  行{row_num} M列: {Path(img_path).name}")

    # 9. 自动调整列宽
    print("\n调整列宽...")
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 9. 保存文件
    output_path = Path(excel_path).parent / f"{Path(excel_path).stem}_已填充.xlsx"
    wb.save(output_path)

    return output_path


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(
        description='将图片表格数据填入Excel模板（增强版）',
        epilog='示例: python map_and_insert.py --excel template.xlsx --image img1.jpg img2.jpg --data data.json --custom \'{"数量":100}\''
    )

    parser.add_argument('--excel', required=True, help='Excel模板文件路径')
    parser.add_argument('--image', nargs='*', help='图片路径（可指定多个）')
    parser.add_argument('--data', help='提取的数据JSON文件路径')
    parser.add_argument('--custom', help='用户自定义内容（JSON格式）')

    args = parser.parse_args()

    # 检查Excel文件
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"错误: Excel文件不存在: {args.excel}")
        sys.exit(1)

    # 解析用户自定义数据
    user_custom = None
    if args.custom:
        try:
            user_custom = json.loads(args.custom)
            print(f"用户自定义数据: {user_custom}")
        except json.JSONDecodeError:
            print(f"错误: 无法解析自定义数据: {args.custom}")
            sys.exit(1)

    # 处理并插入
    print("=" * 60)
    print("图片表格数据填入工具 - 增强版")
    print("=" * 60)

    try:
        output_file = process_and_insert(
            excel_path=str(excel_path),
            image_paths=args.image,
            data_file=args.data,
            user_custom=user_custom
        )

        print("\n" + "=" * 60)
        print("完成!")
        print(f"输出文件: {output_file}")
        print("=" * 60)
    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
