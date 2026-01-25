"""
安全查找Excel最后数据行的工具
用于在插入数据前准确定位插入位置
"""
import sys
from pathlib import Path
from openpyxl import load_workbook


def find_last_data_row_safe(excel_path: str, start_row: int = 5, sheet_name: str = None):
    """
    安全地查找最后一个有数据的行
    
    Args:
        excel_path: Excel文件路径
        start_row: 开始扫描的行号（默认5，跳过表头）
        sheet_name: 工作表名称（默认使用活动工作表）
    
    Returns:
        dict: {
            'last_row': 最后数据行号,
            'insert_row': 建议插入行号,
            'max_row': 工作表max_row,
            'data_rows': 所有有数据的行号列表,
            'last_row_content': 最后一行的内容
        }
    """
    # 打开Excel
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # 获取工作表信息
    max_row = ws.max_row
    
    # 扫描所有行，记录有数据的行
    data_rows = []
    last_data_row = start_row - 1
    
    print(f"开始扫描工作表...")
    print(f"  工作表名称: {ws.title}")
    print(f"  max_row: {max_row}")
    print(f"  扫描范围: 第{start_row}行 - 第{max_row}行")
    print()
    
    # 完整扫描
    for row in range(start_row, max_row + 1):
        # 检查前13列（A-M）是否有数据
        row_has_data = False
        row_content = []
        
        for col in range(1, 14):
            cell_value = ws.cell(row=row, column=col).value
            row_content.append(cell_value)
            
            if cell_value is not None and str(cell_value).strip() != "":
                row_has_data = True
        
        if row_has_data:
            data_rows.append(row)
            last_data_row = row
            
            # 显示有数据的行（前3列）
            preview = [str(v)[:20] if v else "" for v in row_content[:3]]
            print(f"  第{row}行: {' | '.join(preview)}")
    
    # 读取最后一行的完整内容
    last_row_content = {}
    if last_data_row >= start_row:
        for col in range(1, 14):
            cell_value = ws.cell(row=last_data_row, column=col).value
            if cell_value:
                last_row_content[col] = cell_value
    
    # 计算插入位置
    insert_row = last_data_row + 2  # +1为空行，+2为新数据起始行
    
    print()
    print("=" * 60)
    print(f"扫描完成！")
    print(f"  共找到 {len(data_rows)} 行有数据")
    print(f"  最后数据行: 第{last_data_row}行")
    print(f"  建议插入位置: 第{insert_row}行（第{insert_row-1}行为空行分隔）")
    print("=" * 60)
    
    wb.close()
    
    return {
        'last_row': last_data_row,
        'insert_row': insert_row,
        'max_row': max_row,
        'data_rows': data_rows,
        'last_row_content': last_row_content
    }


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python find_last_row.py <excel_path> [start_row]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    start_row = int(sys.argv[2]) if len(sys.argv) > 2 else 5
    
    if not Path(excel_path).exists():
        print(f"错误: 文件不存在: {excel_path}")
        sys.exit(1)
    
    result = find_last_data_row_safe(excel_path, start_row)
    
    print("\n最后一行内容:")
    for col, value in result['last_row_content'].items():
        print(f"  列{col}: {value}")


if __name__ == '__main__':
    main()
