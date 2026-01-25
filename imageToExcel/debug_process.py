"""
Debug the full processing workflow - 完整流程检测
"""
import sys
from pathlib import Path

# Add skill scripts to path
skill_scripts = Path(r"d:\claude code -11\image-table-excel\scripts")
if str(skill_scripts) not in sys.path:
    sys.path.insert(0, str(skill_scripts))

# 导入处理函数
from extract_table_from_image import extract_table_from_image
from map_and_insert import process_and_insert

# 测试文件
image_path = r"d:\claude code -11\imageToExcel\test_image.jpg"
excel_path = r"d:\claude code -11\imageToExcel\中原证券进场材料统计表模板.xlsx"

print("=" * 60)
print("Step 1: OCR识别 - 从图片提取表格数据")
print("=" * 60)
print(f"图片文件: {image_path}")

# 提取数据
try:
    extracted_data = extract_table_from_image(image_path)
    print(f"识别成功: {len(extracted_data)} 行数据")

    for i, row in enumerate(extracted_data):
        print(f"  Row {i+1}: {row}")
except Exception as e:
    print(f"识别失败: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# 保存到JSON
import json
data_file = Path(r"d:\claude code -11\imageToExcel\uploads\extracted_data.json")
data_file.parent.mkdir(exist_ok=True)

with open(data_file, 'w', encoding='utf-8') as f:
    json.dump(extracted_data, f, ensure_ascii=False, indent=2)
print(f"\n数据已保存到: {data_file}")

print("\n" + "=" * 60)
print("Step 2: 数据映射和插入 - 写入Excel")
print("=" * 60)
print(f"Excel模板: {excel_path}")
print(f"图片文件: {image_path}")
print(f"数据文件: {data_file}")

# 处理并插入
try:
    output_path = process_and_insert(
        excel_path=excel_path,
        image_paths=[image_path],
        data_file=str(data_file),
        user_custom=None
    )
    print(f"\n处理成功: {output_path}")
except Exception as e:
    print(f"处理失败: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "=" * 60)
print("Step 3: 结果验证 - 检查生成的Excel")
print("=" * 60)

# 验证结果
from openpyxl import load_workbook
wb = load_workbook(output_path)
ws = wb.active

print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

print("\n最后10行数据:")
for row in range(max(4, ws.max_row-10), ws.max_row+1):
    row_data = []
    for col in range(1, min(14, ws.max_column+1)):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value)[:12] if cell.value else ''
        row_data.append(val)
    print(f"  Row {row}: {' | '.join(row_data)}")

# 检查图片
print("\n检查图片插入:")
image_count = 0
for row in range(ws.max_row-5, ws.max_row+1):
    for col in range(13, 16):
        cell = ws.cell(row=row, column=col)
        if cell.value and ('DISPIMG' in str(cell.value) or isinstance(cell.value, str) and 'image' in cell.value.lower()):
            image_count += 1
            print(f"  发现图片: Row {row}, Column {col}")

print(f"\n图片引用数量: {image_count}")
print("\n✅ 完整流程检测完成!")
