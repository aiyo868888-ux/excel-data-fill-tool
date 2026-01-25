"""
Simple template validator - no unicode characters
"""
import sys
from pathlib import Path

def validate_template(file_path):
    """Validate Excel template"""
    print(f"\nValidating: {file_path}")
    print("=" * 60)

    path = Path(file_path)

    # Check existence
    if not path.exists():
        print("[ERROR] File not found")
        return False

    # Check size
    size = path.stat().st_size
    print(f"[OK] File size: {size} bytes ({size/1024:.2f} KB)")

    # Check extension
    ext = path.suffix.lower()
    print(f"[OK] Extension: {ext}")

    # Try to open with openpyxl
    print("\nTrying to open with openpyxl...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        print("[OK] File can be opened by openpyxl")
        print(f"[OK] Number of worksheets: {len(wb.sheetnames)}")
        print(f"[OK] Worksheet names: {wb.sheetnames}")
        print(f"[OK] Active worksheet: {wb.active.title}")

        # Show some cell data
        ws = wb.active
        print(f"\nFirst 10 rows of active sheet:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(str(cell.value)[:15] if cell.value else "")
            print(f"  Row {row}: {row_data}")

        wb.close()
        print("\n[SUCCESS] Template file is valid!")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to open: {type(e).__name__}: {e}")
        return False

if __name__ == '__main__':
    template_path = r"d:\claude code -11\imageToExcel\中原证券进场材料统计表模板.xlsx"
    validate_template(template_path)
