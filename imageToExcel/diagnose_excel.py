"""
Excel文件诊断工具
帮助诊断为什么openpyxl无法打开Excel文件
"""
import sys
from pathlib import Path

def diagnose_excel_file(file_path):
    """诊断Excel文件"""
    print(f"\n诊断文件: {file_path}")
    print("=" * 60)

    # 1. 检查文件是否存在
    path = Path(file_path)
    if not path.exists():
        print("❌ 文件不存在")
        return

    # 2. 检查文件大小
    size = path.stat().st_size
    print(f"📁 文件大小: {size} bytes ({size/1024:.2f} KB)")

    # 3. 检查文件扩展名
    ext = path.suffix.lower()
    print(f"📝 文件扩展名: {ext}")

    # 4. 检查文件头（魔术数字）
    print("\n检查文件签名...")
    with open(path, 'rb') as f:
        header = f.read(8)
        print(f"   文件头（hex）: {header.hex()}")

        # ZIP文件签名 (xlsx实际是ZIP格式)
        if header[:4] == b'PK\x03\x04':
            print("   ✓ 符合ZIP格式（.xlsx文件）")
        elif header[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A':
            print("   ✓ 符合OLE格式（.xls文件）")
        else:
            print(f"   ✗ 未知的文件格式")
            print(f"   提示: 前8字节: {header}")

    # 5. 尝试用openpyxl打开
    print("\n尝试用openpyxl打开...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        print(f"   ✓ openpyxl可以打开")
        print(f"   工作表数量: {len(wb.sheetnames)}")
        print(f"   工作表名称: {wb.sheetnames}")
        print(f"   活动工作表: {wb.active.title}")
        wb.close()
    except Exception as e:
        print(f"   ✗ openpyxl打开失败: {e}")

    # 6. 尝试读取为文本
    print("\n尝试读取前100字节作为文本...")
    try:
        with open(path, 'rb') as f:
            content = f.read(100)
        print(f"   前100字节: {content[:50]}")
        if b'<?xml' in content or b'<worksheet' in content:
            print("   ✓ 看起来是XML内容")
        if b',' in content and content.count(b',') > 5:
            print("   ✓ 可能是CSV文件")
    except Exception as e:
        print(f"   ✗ 文本读取失败: {e}")

    print("\n" + "=" * 60)
    print("建议:")
    print("1. 用Excel或WPS打开此文件")
    print("2. 如果可以打开，使用'另存为'保存为.xlsx格式")
    print("3. 如果无法打开，文件可能已损坏")
    print("4. 检查文件来源，确保是真正的Excel文件")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python diagnose_excel.py <excel文件路径>")
        print("\n示例:")
        print("  python diagnose_excel.py template.xlsx")
        print("  python diagnose_excel.py ../uploads/test.xlsx")
    else:
        diagnose_excel_file(sys.argv[1])
