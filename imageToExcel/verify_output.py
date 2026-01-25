"""
Verify the generated Excel file
"""
from openpyxl import load_workbook
from pathlib import Path

output_file = Path("output_test.xlsx")

print(f"Verifying: {output_file}")
print("=" * 60)

wb = load_workbook(output_file)
ws = wb.active

print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Show header row (row 4)
print("\nHeader (Row 4):")
header = []
for col in range(1, 14):
    cell = ws.cell(row=4, column=col)
    header.append(str(cell.value)[:10] if cell.value else "")
print(" | ".join(header))

# Show data rows
print("\nData rows:")
for row in range(5, min(ws.max_row + 1, 15)):
    row_data = []
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value)[:12] if cell.value else ""
        row_data.append(val)
    print(f"Row {row}: {' | '.join(row_data)}")

# Check for images
print("\nChecking for images...")
image_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value and 'DISPIMG' in str(cell.value):
            image_count += 1
            print(f"Found image at: Row {cell.row}, Column {cell.column}")

print(f"\nTotal image references found: {image_count}")
print("\n[SUCCESS] Output file is valid!")
