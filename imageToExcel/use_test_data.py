"""
使用测试数据直接测试映射和插入功能
"""
import subprocess
import sys
from pathlib import Path

# 测试文件
excel_path = Path(r"d:\claude code -11\imageToExcel\中原证券进场材料统计表模板.xlsx").absolute()
uploads_dir = Path(r"d:\claude code -11\imageToExcel\uploads")

# 使用image-table-excel自带的测试数据
test_data = Path(r"d:\claude code -11\image-table-excel\extracted_data.json")
data_file = uploads_dir / "extracted_data.json"

# 复制测试数据
import shutil
shutil.copy(test_data, data_file)

print("=" * 80)
print("测试数据映射和插入功能")
print("=" * 80)

# 读取并显示测试数据
import json
with open(data_file, 'r', encoding='utf-8') as f:
    extracted_data = json.load(f)

print(f"\n测试数据 ({len(extracted_data)} 行):")
for i, row in enumerate(extracted_data):
    print(f"  Row {i+1}: {row}")

# 调用map_and_insert脚本
print(f"\n调用 map_and_insert 脚本...")

map_script = Path(r"d:\claude code -11\image-table-excel\scripts\map_and_insert.py")
output_file = uploads_dir / f"{excel_path.stem}_已填充.xlsx"

cmd = [
    sys.executable, str(map_script),
    "--excel", str(excel_path),
    "--data", str(data_file)
]

print(f"命令: {' '.join(cmd)}\n")

result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='ignore')

if result.returncode == 0:
    print("处理成功!")
    if result.stdout:
        print(result.stdout[:500])
else:
    print(f"处理失败! 返回码: {result.returncode}")
    if result.stderr:
        print("错误信息:", result.stderr[:500])
    sys.exit(1)

# 验证结果
print(f"\n验证结果...")
from openpyxl import load_workbook

if output_file.exists():
    wb = load_workbook(output_file)
    ws = wb.active

    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}")

    print(f"\n最后10行:")
    for row in range(max(4, ws.max_row-10), ws.max_row+1):
        row_data = []
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value)[:10] if cell.value else ""
            row_data.append(val)
        print(f"  Row {row}: {' | '.join(row_data)}")

    print("\n✅ 测试完成!")
else:
    print(f"错误: 输出文件未找到 - {output_file}")
