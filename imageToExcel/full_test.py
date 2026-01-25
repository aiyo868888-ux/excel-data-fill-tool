"""
完整流程测试 - 模拟Web应用的处理流程
"""
import subprocess
import json
import sys
from pathlib import Path

# 测试文件
image_path = Path(r"d:\claude code -11\imageToExcel\test_image.jpg").absolute()
excel_path = Path(r"d:\claude code -11\imageToExcel\中原证券进场材料统计表模板.xlsx").absolute()
uploads_dir = Path(r"d:\claude code -11\imageToExcel\uploads")

print("=" * 80)
print("完整流程测试 - 图片表格识别并填入Excel")
print("=" * 80)

print(f"\n[文件信息]")
print(f"  Excel模板: {excel_path.name}")
print(f"  测试图片: {image_path.name}")
print(f"  图片大小: {image_path.stat().st_size / 1024:.1f} KB")

# Step 1: OCR识别
print(f"\n[Step 1] OCR识别 - 从图片提取表格数据")
print("-" * 80)

extract_script = Path(r"d:\claude code -11\image-table-excel\scripts\extract_table_from_image.py")
data_file = uploads_dir / "extracted_data.json"

cmd = [sys.executable, str(extract_script), str(image_path), "--output", str(data_file)]
print(f"命令: {' '.join(cmd)}")

result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='ignore')

if result.returncode == 0:
    print(f"识别成功!")
    if result.stdout:
        print(result.stdout[:500])  # 只打印前500字符
else:
    print(f"识别失败! 返回码: {result.returncode}")
    if result.stderr:
        print("错误信息:", result.stderr[:500])
    sys.exit(1)

# 读取提取的数据
if data_file.exists():
    with open(data_file, 'r', encoding='utf-8') as f:
        extracted_data = json.load(f)

    print(f"\n提取的数据 ({len(extracted_data)} 行):")
    for i, row in enumerate(extracted_data[:5]):
        print(f"  Row {i+1}: {row}")
else:
    print(f"错误: 数据文件未生成 - {data_file}")
    sys.exit(1)

# Step 2: 数据映射和插入
print(f"\n[Step 2] 数据映射和插入 - 写入Excel")
print("-" * 80)

map_script = Path(r"d:\claude code -11\image-table-excel\scripts\map_and_insert.py")
output_file = uploads_dir / f"{excel_path.stem}_已填充.xlsx"

cmd = [
    sys.executable, str(map_script),
    "--excel", str(excel_path),
    "--image", str(image_path),
    "--data", str(data_file)
]
print(f"命令: {' '.join(cmd)}")

result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='ignore')

if result.returncode == 0:
    print(f"处理成功!")
    if result.stdout:
        print(result.stdout[:500])

    # 查找生成的文件
    if output_file.exists():
        print(f"输出文件: {output_file}")
    else:
        # 可能在父目录
        output_file = excel_path.parent / f"{excel_path.stem}_已填充.xlsx"
        if output_file.exists():
            print(f"输出文件: {output_file}")
        else:
            print(f"警告: 未找到输出文件")
            sys.exit(1)
else:
    print(f"处理失败! 返回码: {result.returncode}")
    if result.stderr:
        print("错误信息:", result.stderr[:500])
    sys.exit(1)

# Step 3: 验证结果
print(f"\n[Step 3] 结果验证 - 检查生成的Excel")
print("-" * 80)

from openpyxl import load_workbook

wb = load_workbook(output_file)
ws = wb.active

print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

# 显示表头
print(f"\n表头 (Row 4):")
header = []
for col in range(1, 14):
    cell = ws.cell(row=4, column=col)
    header.append(str(cell.value)[:8] if cell.value else "")
print("  " + " | ".join(header))

# 显示所有数据行
print(f"\n所有数据行:")
for row in range(5, ws.max_row + 1):
    row_data = []
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        val = str(cell.value)[:10] if cell.value else ""
        row_data.append(val)

    # 只显示有数据的行
    if any(row_data):
        print(f"  Row {row}: {' | '.join(row_data)}")

# 检查图片
print(f"\n检查图片插入:")
image_count = 0
for row_idx in range(5, ws.max_row + 1):
    for col_idx in range(13, 16):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and ('DISPIMG' in str(cell.value) or 'image' in str(cell.value).lower()):
            image_count += 1
            col_letter = chr(64 + col_idx)
            print(f"  发现图片: Row {row_idx}, Column {col_idx} ({col_letter}列)")

print(f"\n图片引用数量: {image_count}")

# 总结
print("\n" + "=" * 80)
print("测试总结")
print("=" * 80)

original_rows = 8  # 模板原有行数
new_rows = ws.max_row - original_rows
print(f"  原有行数: {original_rows}")
print(f"  当前行数: {ws.max_row}")
print(f"  新增行数: {new_rows}")
print(f"  提取数据: {len(extracted_data)} 行")
print(f"  插入图片: {image_count} 个")

if new_rows > 0:
    print("\n✅ 测试成功! 数据已正确插入")
else:
    print("\n⚠️  警告: 似乎没有新数据插入")
