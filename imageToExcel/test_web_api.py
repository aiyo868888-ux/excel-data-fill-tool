"""
测试Web API - 使用OCR识别
"""
import requests
import time

url = "http://localhost:5000/api/process"

files = {
    'excel_file': open(r'd:\claude code -11\imageToExcel\中原证券进场材料统计表模板.xlsx', 'rb'),
    'image_files': open(r'd:\claude code -11\imageToExcel\test_image.jpg', 'rb'),
}

print("上传文件到Web API...")
print(f"  Excel: 中原证券进场材料统计表模板.xlsx")
print(f"  Image: test_image.jpg")
print()

try:
    start = time.time()
    response = requests.post(url, files=files, timeout=120)
    elapsed = time.time() - start

    print(f"响应时间: {elapsed:.1f}秒")
    print(f"状态码: {response.status_code}")
    print()

    result = response.json()

    if result.get('success'):
        print("处理成功!")
        print(f"  输出文件: {result.get('output_file')}")
        print(f"  下载链接: {result.get('download_url')}")
        print(f"  数据: {result.get('data')}")

        # 下载文件
        download_url = "http://localhost:5000" + result['download_url']
        print(f"\n正在下载: {download_url}")

        result_response = requests.get(download_url)
        output_file = "web_api_output.xlsx"

        with open(output_file, 'wb') as f:
            f.write(result_response.content)

        print(f"已保存: {output_file}")

        # 验证
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active

        print(f"\n输出文件验证:")
        print(f"  总行数: {ws.max_row}")
        print(f"\n  最后5行数据:")

        for row in range(max(4, ws.max_row-5), ws.max_row+1):
            row_data = []
            for col in [1, 4, 5, 6, 7, 12]:
                cell = ws.cell(row=row, column=col)
                val = str(cell.value)[:8] if cell.value else ''
                row_data.append(val)

            if any(row_data):
                marker = " [OCR识别]" if row > 9 else ""
                print(f"    Row {row}{marker}: {', '.join(row_data)}")

    else:
        print("处理失败!")
        print(f"  错误: {result.get('error')}")

except Exception as e:
    print(f"异常: {e}")
    import traceback
    traceback.print_exc()

finally:
    for f in files.values():
        f.close()
